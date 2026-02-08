from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Company, AuditLog, UserActivitySummary
from .forms import CompanyForm
from .audit_utils import log_export


@login_required
def company_list(request):
    """List all companies (Super Admin only)"""
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can access company management')
        return redirect('assets:dashboard')
    
    companies = Company.objects.filter(is_deleted=False).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(companies, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'core/company_list.html', context)


@login_required
def company_create(request):
    """Create new company (Super Admin only)"""
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can create companies')
        return redirect('assets:dashboard')
    
    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES)
        if form.is_valid():
            company = form.save()
            messages.success(request, f'Company "{company.name}" created successfully!')
            return redirect('core:company_list')
    else:
        form = CompanyForm()
    
    context = {
        'form': form,
        'action': 'Create',
    }
    
    return render(request, 'core/company_form.html', context)


@login_required
def company_detail(request, pk):
    """View company details (Super Admin only)"""
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can view company details')
        return redirect('assets:dashboard')
    
    company = get_object_or_404(Company, pk=pk, is_deleted=False)
    
    # Get company statistics
    from assets.models import Asset
    from users.models import UserProfile
    
    stats = {
        'total_assets': Asset.objects.filter(company=company, is_deleted=False).count(),
        'total_users': UserProfile.objects.filter(company=company).count(),
        'active_assets': Asset.objects.filter(company=company, status='IN_USE', is_deleted=False).count(),
    }
    
    context = {
        'company': company,
        'stats': stats,
    }
    
    return render(request, 'core/company_detail.html', context)


@login_required
def company_update(request, pk):
    """Update company (Super Admin only)"""
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can update companies')
        return redirect('assets:dashboard')
    
    company = get_object_or_404(Company, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            company = form.save()
            messages.success(request, f'Company "{company.name}" updated successfully!')
            return redirect('core:company_detail', pk=company.pk)
    else:
        form = CompanyForm(instance=company)
    
    context = {
        'form': form,
        'company': company,
        'action': 'Update',
    }
    
    return render(request, 'core/company_form.html', context)


@login_required
def company_delete(request, pk):
    """Delete company (Super Admin only)"""
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can delete companies')
        return redirect('assets:dashboard')
    
    company = get_object_or_404(Company, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        company.is_deleted = True
        company.save()
        messages.success(request, f'Company "{company.name}" deleted successfully!')
        return redirect('core:company_list')
    
    context = {
        'company': company,
    }
    
    return render(request, 'core/company_confirm_delete.html', context)


@login_required
def set_company_context(request, company_id=None):
    """Set company context for super admin (allows them to filter by company)"""
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can switch company context')
        return redirect('assets:dashboard')
    
    if company_id:
        company = get_object_or_404(Company, pk=company_id, is_deleted=False)
        request.session['selected_company_id'] = company.id
        messages.success(request, f'Now viewing data for: {company.name}')
    else:
        # Clear company context (view all companies)
        request.session.pop('selected_company_id', None)
        messages.success(request, 'Now viewing data for all companies')
    
    # Redirect to the page they came from or dashboard
    next_url = request.GET.get('next', reverse('assets:dashboard'))
    return redirect(next_url)


def is_admin(user):
    """Check if user is super admin or company admin"""
    return user.is_superuser or (hasattr(user, 'profile') and user.profile.is_company_admin)


@login_required
@user_passes_test(is_admin)
def audit_log_list(request):
    """View audit logs with filtering"""
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Base queryset
    if company:
        logs = AuditLog.objects.filter(company=company).select_related('user', 'content_type', 'company')
    elif is_super_admin:
        logs = AuditLog.objects.all().select_related('user', 'content_type', 'company')
    else:
        # Company admin
        try:
            company = request.user.profile.company
            logs = AuditLog.objects.filter(company=company).select_related('user', 'content_type', 'company')
        except:
            logs = AuditLog.objects.none()
    
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            date_from = ''
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            date_to = ''
    
    # Action filter
    action = request.GET.get('action', '')
    if action:
        logs = logs.filter(action=action)
    
    # User filter
    user_id = request.GET.get('user', '')
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    # Model filter
    model = request.GET.get('model', '')
    if model:
        logs = logs.filter(content_type__model=model)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        logs = logs.filter(
            Q(description__icontains=search) |
            Q(username__icontains=search) |
            Q(object_repr__icontains=search)
        )
    
    # Company filter (super admin only)
    company_filter = request.GET.get('company', '')
    if company_filter and is_super_admin:
        logs = logs.filter(company_id=company_filter)
    
    logs = logs.order_by('-timestamp')
    
    # Pagination
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    from django.contrib.auth.models import User
    users = User.objects.filter(is_active=True).order_by('username')
    
    from django.contrib.contenttypes.models import ContentType
    models = ContentType.objects.filter(
        id__in=AuditLog.objects.values_list('content_type_id', flat=True).distinct()
    ).order_by('model')
    
    companies = None
    if is_super_admin:
        companies = Company.objects.filter(is_deleted=False, is_active=True)
    
    context = {
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'action': action,
        'user_id': user_id,
        'model': model,
        'search': search,
        'company_filter': company_filter,
        'users': users,
        'models': models,
        'action_choices': AuditLog.ACTION_CHOICES,
        'company': company,
        'is_super_admin': is_super_admin,
        'companies': companies,
    }
    
    return render(request, 'core/audit_log_list.html', context)


@login_required
@user_passes_test(is_admin)
def audit_log_detail(request, pk):
    """View detailed audit log entry"""
    log = get_object_or_404(AuditLog, pk=pk)
    
    # Check permission
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if not is_super_admin:
        if company and log.company != company:
            messages.error(request, 'You do not have permission to view this audit log.')
            return redirect('core:audit_log_list')
        elif not company:
            try:
                if log.company != request.user.profile.company:
                    messages.error(request, 'You do not have permission to view this audit log.')
                    return redirect('core:audit_log_list')
            except:
                pass
    
    context = {
        'log': log,
        'company': company,
        'is_super_admin': is_super_admin,
    }
    
    return render(request, 'core/audit_log_detail.html', context)


@login_required
@user_passes_test(is_admin)
def audit_log_export(request):
    """Export audit logs to Excel"""
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Get filters from GET parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    action = request.GET.get('action', '')
    user_id = request.GET.get('user', '')
    model = request.GET.get('model', '')
    search = request.GET.get('search', '')
    company_filter = request.GET.get('company', '')
    
    # Base queryset (same logic as list view)
    if company:
        logs = AuditLog.objects.filter(company=company).select_related('user', 'content_type', 'company')
    elif is_super_admin:
        logs = AuditLog.objects.all().select_related('user', 'content_type', 'company')
    else:
        try:
            company = request.user.profile.company
            logs = AuditLog.objects.filter(company=company).select_related('user', 'content_type', 'company')
        except:
            logs = AuditLog.objects.none()
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass
    
    if action:
        logs = logs.filter(action=action)
    
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    if model:
        logs = logs.filter(content_type__model=model)
    
    if search:
        logs = logs.filter(
            Q(description__icontains=search) |
            Q(username__icontains=search) |
            Q(object_repr__icontains=search)
        )
    
    if company_filter and is_super_admin:
        logs = logs.filter(company_id=company_filter)
    
    logs = logs.order_by('-timestamp')[:5000]  # Limit to 5000 records
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C17845", end_color="C17845", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Timestamp', 'User', 'Action', 'Object Type', 'Object', 'Description', 
               'Changed Fields', 'IP Address', 'Company']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, log in enumerate(logs, 2):
        ws.cell(row=row_num, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=2, value=log.username)
        ws.cell(row=row_num, column=3, value=log.action_display)
        ws.cell(row=row_num, column=4, value=log.content_type.model if log.content_type else '-')
        ws.cell(row=row_num, column=5, value=log.object_repr)
        ws.cell(row=row_num, column=6, value=log.description)
        ws.cell(row=row_num, column=7, value=', '.join(log.changed_fields) if log.changed_fields else '-')
        ws.cell(row=row_num, column=8, value=log.ip_address or '-')
        ws.cell(row=row_num, column=9, value=log.company.name if log.company else '-')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=audit_logs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    
    # Log the export action
    log_export(request, 'AuditLog', logs.count(), 'Excel')
    
    return response


@login_required
@user_passes_test(is_admin)
def user_activity_report(request):
    """View user activity summary report"""
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Date range (default to last 30 days)
    date_to = timezone.now().date()
    date_from = date_to - timedelta(days=30)
    
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Base queryset
    if company:
        logs = AuditLog.objects.filter(company=company, timestamp__date__gte=date_from, timestamp__date__lte=date_to)
    elif is_super_admin:
        logs = AuditLog.objects.filter(timestamp__date__gte=date_from, timestamp__date__lte=date_to)
    else:
        try:
            company = request.user.profile.company
            logs = AuditLog.objects.filter(company=company, timestamp__date__gte=date_from, timestamp__date__lte=date_to)
        except:
            logs = AuditLog.objects.none()
    
    # Aggregate by user
    from django.contrib.auth.models import User
    user_stats = logs.values('user', 'username').annotate(
        total_actions=Count('id'),
        creates=Count('id', filter=Q(action='CREATE')),
        updates=Count('id', filter=Q(action='UPDATE')),
        deletes=Count('id', filter=Q(action='DELETE')),
        views=Count('id', filter=Q(action='VIEW')),
        exports=Count('id', filter=Q(action='EXPORT')),
        logins=Count('id', filter=Q(action='LOGIN')),
    ).order_by('-total_actions')
    
    # Overall statistics
    total_actions = logs.count()
    total_creates = logs.filter(action='CREATE').count()
    total_updates = logs.filter(action='UPDATE').count()
    total_deletes = logs.filter(action='DELETE').count()
    total_views = logs.filter(action='VIEW').count()
    total_exports = logs.filter(action='EXPORT').count()
    
    # Most active users (top 10)
    top_users = user_stats[:10]
    
    # Activity by day
    from django.db.models.functions import TruncDate
    daily_activity = logs.annotate(
        date=TruncDate('timestamp')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'user_stats': user_stats,
        'top_users': top_users,
        'daily_activity': daily_activity,
        'total_actions': total_actions,
        'total_creates': total_creates,
        'total_updates': total_updates,
        'total_deletes': total_deletes,
        'total_views': total_views,
        'total_exports': total_exports,
        'company': company,
        'is_super_admin': is_super_admin,
    }
    
    return render(request, 'core/user_activity_report.html', context)


@login_required
@user_passes_test(is_admin)
def user_activity_export(request):
    """Export user activity report to Excel"""
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Date range
    date_to = timezone.now().date()
    date_from = date_to - timedelta(days=30)
    
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Base queryset
    if company:
        logs = AuditLog.objects.filter(company=company, timestamp__date__gte=date_from, timestamp__date__lte=date_to)
    elif is_super_admin:
        logs = AuditLog.objects.filter(timestamp__date__gte=date_from, timestamp__date__lte=date_to)
    else:
        try:
            company = request.user.profile.company
            logs = AuditLog.objects.filter(company=company, timestamp__date__gte=date_from, timestamp__date__lte=date_to)
        except:
            logs = AuditLog.objects.none()
    
    # Aggregate by user
    user_stats = logs.values('user', 'username').annotate(
        total_actions=Count('id'),
        creates=Count('id', filter=Q(action='CREATE')),
        updates=Count('id', filter=Q(action='UPDATE')),
        deletes=Count('id', filter=Q(action='DELETE')),
        views=Count('id', filter=Q(action='VIEW')),
        exports=Count('id', filter=Q(action='EXPORT')),
        logins=Count('id', filter=Q(action='LOGIN')),
    ).order_by('-total_actions')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "User Activity"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C17845", end_color="C17845", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['User', 'Total Actions', 'Creates', 'Updates', 'Deletes', 'Views', 'Exports', 'Logins']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, stat in enumerate(user_stats, 2):
        ws.cell(row=row_num, column=1, value=stat['username'])
        ws.cell(row=row_num, column=2, value=stat['total_actions'])
        ws.cell(row=row_num, column=3, value=stat['creates'])
        ws.cell(row=row_num, column=4, value=stat['updates'])
        ws.cell(row=row_num, column=5, value=stat['deletes'])
        ws.cell(row=row_num, column=6, value=stat['views'])
        ws.cell(row=row_num, column=7, value=stat['exports'])
        ws.cell(row=row_num, column=8, value=stat['logins'])
    
    # Adjust column widths
    for i in range(1, 9):
        ws.column_dimensions[chr(64+i)].width = 15
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=user_activity_{date_from}_{date_to}.xlsx'
    
    wb.save(response)
    
    # Log the export
    log_export(request, 'UserActivity', len(user_stats), 'Excel')
    
    return response
