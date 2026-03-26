"""
Excel Import Utility for Assets
Handles importing asset data from Excel spreadsheets
"""
from openpyxl import load_workbook
from django.contrib.auth.models import User
from django.db.models import Q
from datetime import datetime
from decimal import Decimal
from .models import Asset, AssetCategory, AssetType, Vendor
from users.models import Location, Department


class AssetExcelImporter:
    """Handles importing assets from Excel files"""
    
    def __init__(self, file, company, user):
        self.file = file
        self.company = company
        self.user = user
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.skip_count = 0
    
    def parse_date(self, value):
        """Parse date from various formats"""
        if not value:
            return None
        
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, str):
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except:
                    continue
        
        return None
    
    def parse_decimal(self, value):
        """Parse decimal from string or number"""
        if not value:
            return None
        
        try:
            if isinstance(value, str):
                # Remove currency symbols and commas
                value = value.replace('K', '').replace('$', '').replace(',', '').strip()
            return Decimal(str(value))
        except:
            return None
    
    def get_or_none(self, model, company, **kwargs):
        """Get object or return None"""
        try:
            filters = {'company': company, 'is_deleted': False}
            filters.update(kwargs)
            return model.objects.get(**filters)
        except model.DoesNotExist:
            return None
        except model.MultipleObjectsReturned:
            return model.objects.filter(**filters).first()
    
    def import_assets(self):
        """Import assets from Excel file"""
        try:
            wb = load_workbook(self.file, data_only=True)
            ws = wb.active
            
            # Expected columns (row 1)
            headers = [cell.value for cell in ws[1]]
            
            # Define expected column mapping
            expected_columns = {
                'asset_tag':             ['Asset Tag', 'Tag', 'Asset ID'],
                'name':                  ['Asset Name', 'Name'],
                'description':           ['Description'],
                'category':              ['Category', 'Asset Category'],
                'asset_type':            ['Asset Type', 'Type'],
                'make':                  ['Make / Manufacturer', 'Make', 'Manufacturer'],
                'model':                 ['Model'],
                'serial_number':         ['Serial Number', 'Serial No', 'S/N'],
                'status':                ['Status'],
                'condition':             ['Condition'],
                'location':              ['Location'],
                'department':            ['Department'],
                'assigned_to':           ['Assigned To', 'Assigned User', 'User'],
                'vendor':                ['Vendor'],
                'purchase_date':         ['Purchase Date', 'PO Date'],
                'purchase_price':        ['Purchase Price', 'Price', 'Cost'],
                'purchase_order_number': ['Purchase Order No', 'PO Number', 'PO No'],
                'invoice_number':        ['Invoice Number', 'Invoice No'],
                'invoice_date':          ['Invoice Date'],
                'warranty_start_date':   ['Warranty Start Date'],
                'warranty_end_date':     ['Warranty End Date', 'Warranty Expiry'],
                'warranty_period_months':['Warranty Months', 'Warranty Period Months'],
                'amc_start_date':        ['AMC Start Date'],
                'amc_end_date':          ['AMC End Date'],
                'amc_cost':              ['AMC Cost'],
                'depreciation_rate':     ['Depreciation Rate %', 'Depreciation Rate'],
                'useful_life_years':     ['Useful Life Years', 'Useful Life'],
                'salvage_value':         ['Salvage Value'],
                'notes':                 ['Notes'],
                'is_critical':           ['Is Critical (Y/N)', 'Is Critical'],
                'is_insured':            ['Is Insured (Y/N)', 'Is Insured'],
                'insurance_policy_number':['Insurance Policy No', 'Insurance Policy Number'],
                'insurance_expiry_date': ['Insurance Expiry', 'Insurance Expiry Date'],
            }
            
            # Create column mapping
            col_map = {}
            for field, possible_names in expected_columns.items():
                for idx, header in enumerate(headers):
                    if header and header.strip() in possible_names:
                        col_map[field] = idx
                        break
            
            # Process each row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Extract data
                    asset_tag = row[col_map.get('asset_tag', 0)] if 'asset_tag' in col_map else None
                    if not asset_tag:
                        self.errors.append(f"Row {row_idx}: Asset Tag is required")
                        self.skip_count += 1
                        continue
                    
                    # Check if asset already exists
                    if Asset.objects.filter(company=self.company, asset_tag=asset_tag, is_deleted=False).exists():
                        self.warnings.append(f"Row {row_idx}: Asset {asset_tag} already exists - skipped")
                        self.skip_count += 1
                        continue
                    
                    # Get or create related objects
                    category_name = row[col_map.get('category')] if 'category' in col_map else None
                    category = None
                    if category_name:
                        category = self.get_or_none(AssetCategory, self.company, name=category_name)
                        if not category:
                            self.warnings.append(f"Row {row_idx}: Category '{category_name}' not found - using default")
                    
                    if not category:
                        # Use first available category
                        category = AssetCategory.objects.filter(
                            company=self.company, is_deleted=False
                        ).first()
                        if not category:
                            self.errors.append(f"Row {row_idx}: No categories available in company")
                            self.skip_count += 1
                            continue
                    
                    asset_type_name = row[col_map.get('asset_type')] if 'asset_type' in col_map else None
                    asset_type = None
                    if asset_type_name:
                        asset_type = self.get_or_none(AssetType, self.company, name=asset_type_name)
                    
                    if not asset_type:
                        # Use first available type for the category
                        asset_type = AssetType.objects.filter(
                            company=self.company, category=category, is_deleted=False
                        ).first()
                        if not asset_type:
                            self.warnings.append(f"Row {row_idx}: No asset type found - using first available")
                            asset_type = AssetType.objects.filter(
                                company=self.company, is_deleted=False
                            ).first()
                    
                    if not asset_type:
                        self.errors.append(f"Row {row_idx}: No asset types available in company")
                        self.skip_count += 1
                        continue
                    
                    # Create asset
                    asset = Asset(
                        company=self.company,
                        asset_tag=str(asset_tag).strip(),
                        name=row[col_map.get('name', 1)] if 'name' in col_map else f"Asset {asset_tag}",
                        category=category,
                        asset_type=asset_type,
                    )
                    
                    # Optional fields
                    if 'description' in col_map and row[col_map['description']]:
                        asset.description = str(row[col_map['description']])
                    
                    if 'make' in col_map and row[col_map['make']]:
                        asset.make = str(row[col_map['make']])
                    
                    if 'model' in col_map and row[col_map['model']]:
                        asset.model = str(row[col_map['model']])
                    
                    if 'serial_number' in col_map and row[col_map['serial_number']]:
                        serial = str(row[col_map['serial_number']]).strip()
                        # Check for duplicate serial number
                        if Asset.objects.filter(serial_number=serial, is_deleted=False).exists():
                            self.warnings.append(f"Row {row_idx}: Serial number {serial} already exists - left blank")
                        else:
                            asset.serial_number = serial
                    
                    if 'status' in col_map and row[col_map['status']]:
                        status = str(row[col_map['status']]).strip().upper().replace(' ', '_')
                        valid_statuses = [s[0] for s in Asset.STATUS_CHOICES]
                        if status in valid_statuses:
                            asset.status = status
                    
                    if 'condition' in col_map and row[col_map['condition']]:
                        condition = str(row[col_map['condition']]).strip().upper().replace(' ', '_')
                        valid_conditions = [c[0] for c in Asset.CONDITION_CHOICES]
                        if condition in valid_conditions:
                            asset.condition = condition
                    
                    # Location
                    if 'location' in col_map and row[col_map['location']]:
                        location = self.get_or_none(Location, self.company, name=str(row[col_map['location']]))
                        if location:
                            asset.location = location
                    
                    # Department
                    if 'department' in col_map and row[col_map['department']]:
                        department = self.get_or_none(Department, self.company, name=str(row[col_map['department']]))
                        if department:
                            asset.department = department
                    
                    # Assigned user (by username or email)
                    if 'assigned_to' in col_map and row[col_map['assigned_to']]:
                        user_identifier = str(row[col_map['assigned_to']]).strip()
                        try:
                            assigned_user = User.objects.filter(
                                Q(username=user_identifier) | Q(email=user_identifier),
                                profile__company=self.company,
                                is_active=True
                            ).first()
                            if assigned_user:
                                asset.assigned_to = assigned_user
                        except:
                            pass
                    
                    # Vendor
                    if 'vendor' in col_map and row[col_map['vendor']]:
                        vendor = self.get_or_none(Vendor, self.company, name=str(row[col_map['vendor']]))
                        if vendor:
                            asset.vendor = vendor

                    # Financial fields
                    if 'purchase_date' in col_map:
                        asset.purchase_date = self.parse_date(row[col_map['purchase_date']])

                    if 'purchase_price' in col_map:
                        asset.purchase_price = self.parse_decimal(row[col_map['purchase_price']])

                    if 'purchase_order_number' in col_map and row[col_map['purchase_order_number']]:
                        asset.purchase_order_number = str(row[col_map['purchase_order_number']])

                    if 'invoice_number' in col_map and row[col_map['invoice_number']]:
                        asset.invoice_number = str(row[col_map['invoice_number']])

                    if 'invoice_date' in col_map:
                        asset.invoice_date = self.parse_date(row[col_map['invoice_date']])

                    # Warranty fields
                    if 'warranty_start_date' in col_map:
                        asset.warranty_start_date = self.parse_date(row[col_map['warranty_start_date']])

                    if 'warranty_end_date' in col_map:
                        asset.warranty_end_date = self.parse_date(row[col_map['warranty_end_date']])

                    if 'warranty_period_months' in col_map and row[col_map['warranty_period_months']]:
                        try:
                            asset.warranty_period_months = int(row[col_map['warranty_period_months']])
                        except (ValueError, TypeError):
                            pass

                    # AMC fields
                    if 'amc_start_date' in col_map:
                        asset.amc_start_date = self.parse_date(row[col_map['amc_start_date']])

                    if 'amc_end_date' in col_map:
                        asset.amc_end_date = self.parse_date(row[col_map['amc_end_date']])

                    if 'amc_cost' in col_map:
                        asset.amc_cost = self.parse_decimal(row[col_map['amc_cost']])

                    # Depreciation fields
                    if 'depreciation_rate' in col_map:
                        asset.depreciation_rate = self.parse_decimal(row[col_map['depreciation_rate']])

                    if 'useful_life_years' in col_map and row[col_map['useful_life_years']]:
                        try:
                            asset.useful_life_years = int(row[col_map['useful_life_years']])
                        except (ValueError, TypeError):
                            pass

                    if 'salvage_value' in col_map:
                        asset.salvage_value = self.parse_decimal(row[col_map['salvage_value']])

                    # Other fields
                    if 'notes' in col_map and row[col_map['notes']]:
                        asset.notes = str(row[col_map['notes']])

                    if 'is_critical' in col_map and row[col_map['is_critical']]:
                        asset.is_critical = str(row[col_map['is_critical']]).strip().upper() in ('Y', 'YES', 'TRUE', '1')

                    if 'is_insured' in col_map and row[col_map['is_insured']]:
                        asset.is_insured = str(row[col_map['is_insured']]).strip().upper() in ('Y', 'YES', 'TRUE', '1')

                    if 'insurance_policy_number' in col_map and row[col_map['insurance_policy_number']]:
                        asset.insurance_policy_number = str(row[col_map['insurance_policy_number']])

                    if 'insurance_expiry_date' in col_map:
                        asset.insurance_expiry_date = self.parse_date(row[col_map['insurance_expiry_date']])
                    
                    # Save asset
                    asset.save()
                    
                    # Create history entry
                    from .models import AssetHistory
                    AssetHistory.objects.create(
                        asset=asset,
                        action_type='CREATED',
                        performed_by=self.user,
                        remarks=f'Asset imported from Excel'
                    )
                    
                    self.success_count += 1
                    
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
                    self.skip_count += 1
            
            return {
                'success': self.success_count > 0,
                'success_count': self.success_count,
                'skip_count': self.skip_count,
                'errors': self.errors,
                'warnings': self.warnings,
            }
            
        except Exception as e:
            return {
                'success': False,
                'success_count': 0,
                'skip_count': 0,
                'errors': [f"File processing error: {str(e)}"],
                'warnings': [],
            }


def generate_import_template(company=None):
    """
    Generate a comprehensive Excel template for asset import.
    When company is provided, pre-populates reference data (categories, types, locations, departments).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from io import BytesIO

    wb = Workbook()

    # ── Colour palette ───────────────────────────────────────────────
    BRAND      = "C17845"   # primary orange
    BRAND_DARK = "9E5A2B"
    REQ_FILL   = "FFF3E0"   # light orange  – required columns
    OPT_FILL   = "E8F5E9"   # light green   – optional columns
    FIN_FILL   = "E3F2FD"   # light blue    – financial columns
    WARN_FILL  = "FFF9C4"   # light yellow  – depreciation columns
    HEADER_FG  = "FFFFFF"
    SECTION_FG = "FFFFFF"

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def hfill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # ── Fetch company reference data ──────────────────────────────────
    categories, asset_types, locations, departments, vendors = [], [], [], [], []
    if company:
        from .models import AssetCategory, AssetType, Vendor
        from users.models import Location as Loc, Department as Dept
        categories  = list(AssetCategory.objects.filter(company=company, is_deleted=False, is_active=True).values_list('name', flat=True).order_by('name'))
        asset_types = list(AssetType.objects.filter(company=company, is_deleted=False, is_active=True).values_list('name', flat=True).order_by('name'))
        locations   = list(Loc.objects.filter(company=company, is_deleted=False, is_active=True).values_list('name', flat=True).order_by('name'))
        departments = list(Dept.objects.filter(company=company, is_deleted=False, is_active=True).values_list('name', flat=True).order_by('name'))
        vendors     = list(Vendor.objects.filter(company=company, is_deleted=False, is_active=True).values_list('name', flat=True).order_by('name'))

    statuses   = ['PLANNING', 'ORDERED', 'IN_STOCK', 'DEPLOYED', 'IN_USE', 'UNDER_MAINTENANCE', 'RETIRED', 'DISPOSED', 'LOST', 'STOLEN']
    conditions = ['EXCELLENT', 'GOOD', 'FAIR', 'POOR', 'NOT_WORKING']

    # ════════════════════════════════════════════════════════════════
    # SHEET 1 – Import Template
    # ════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Import Template"
    ws.freeze_panes = "A3"   # freeze header rows

    # Column definitions: (header, fill_key, width, hint)
    COLS = [
        # ── Required ──────────────────────────────────────────────────
        ("Asset Tag *",         "req", 16,  "e.g. AST-001  (must be unique)"),
        ("Asset Name *",        "req", 24,  "Full descriptive name"),
        ("Category *",          "req", 20,  "Must match Reference sheet exactly"),
        ("Asset Type *",        "req", 20,  "Must match Reference sheet exactly"),
        # ── Basic Info ────────────────────────────────────────────────
        ("Description",         "opt", 30,  "Optional – detailed description"),
        ("Make / Manufacturer", "opt", 18,  "e.g. Dell, HP, Samsung"),
        ("Model",               "opt", 18,  "e.g. Latitude 5520"),
        ("Serial Number",       "opt", 18,  "Must be globally unique"),
        ("Status",              "opt", 18,  "See valid values in col header"),
        ("Condition",           "opt", 16,  "See valid values in col header"),
        # ── Assignment ────────────────────────────────────────────────
        ("Location",            "opt", 20,  "Must match Reference sheet exactly"),
        ("Department",          "opt", 20,  "Must match Reference sheet exactly"),
        ("Assigned To",         "opt", 22,  "Username or email of user"),
        # ── Procurement ───────────────────────────────────────────────
        ("Vendor",              "opt", 20,  "Must match Reference sheet exactly"),
        ("Purchase Date",       "fin", 16,  "DD/MM/YYYY or YYYY-MM-DD"),
        ("Purchase Price",      "fin", 16,  "Numeric, e.g. 1500.00"),
        ("Purchase Order No",   "fin", 18,  "PO reference number"),
        ("Invoice Number",      "fin", 16,  "Invoice reference"),
        ("Invoice Date",        "fin", 16,  "DD/MM/YYYY or YYYY-MM-DD"),
        # ── Warranty & AMC ────────────────────────────────────────────
        ("Warranty Start Date", "opt", 18,  "DD/MM/YYYY or YYYY-MM-DD"),
        ("Warranty End Date",   "opt", 18,  "DD/MM/YYYY or YYYY-MM-DD"),
        ("Warranty Months",     "opt", 16,  "Number, e.g. 24"),
        ("AMC Start Date",      "opt", 16,  "DD/MM/YYYY or YYYY-MM-DD"),
        ("AMC End Date",        "opt", 16,  "DD/MM/YYYY or YYYY-MM-DD"),
        ("AMC Cost",            "fin", 14,  "Numeric, e.g. 500.00"),
        # ── Depreciation ──────────────────────────────────────────────
        ("Depreciation Rate %", "dep", 18,  "Annual rate, e.g. 20 (for 20%)"),
        ("Useful Life Years",   "dep", 18,  "Integer, e.g. 5"),
        ("Salvage Value",       "dep", 16,  "Numeric, e.g. 100.00"),
        # ── Other ─────────────────────────────────────────────────────
        ("Notes",               "opt", 30,  "Any additional notes"),
        ("Is Critical (Y/N)",   "opt", 16,  "Y or N"),
        ("Is Insured (Y/N)",    "opt", 16,  "Y or N"),
        ("Insurance Policy No", "opt", 20,  "Policy reference number"),
        ("Insurance Expiry",    "opt", 16,  "DD/MM/YYYY or YYYY-MM-DD"),
    ]

    fill_map = {
        "req": hfill(BRAND),
        "opt": hfill("2E7D32"),
        "fin": hfill("1565C0"),
        "dep": hfill("E65100"),
    }
    hint_fill_map = {
        "req": hfill(REQ_FILL),
        "opt": hfill(OPT_FILL),
        "fin": hfill(FIN_FILL),
        "dep": hfill(WARN_FILL),
    }

    # Row 1 – section labels
    section_spans = [
        (1,  4,  "REQUIRED",      BRAND),
        (5,  13, "BASIC INFO",    "2E7D32"),
        (14, 19, "PROCUREMENT",   "1565C0"),
        (20, 25, "WARRANTY & AMC","00695C"),
        (26, 28, "DEPRECIATION",  "E65100"),
        (29, 34, "OTHER",         "4A148C"),
    ]
    for start_col, end_col, label, color in section_spans:
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.fill = hfill(color)
        cell.font = Font(bold=True, color=HEADER_FG, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    ws.row_dimensions[1].height = 18

    # Row 2 – column headers
    for col_idx, (header, fill_key, width, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = fill_map[fill_key]
        cell.font = Font(bold=True, color=HEADER_FG, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[2].height = 30

    # Row 3 – hints / notes
    for col_idx, (_, fill_key, _, hint) in enumerate(COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hint)
        cell.fill = hint_fill_map[fill_key]
        cell.font = Font(italic=True, size=8, color="555555")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[3].height = 22

    # ── Example rows (4 & 5) ─────────────────────────────────────────
    example_rows = [
        [
            'AST-001', 'Dell Laptop 15"', categories[0] if categories else 'IT Equipment',
            asset_types[0] if asset_types else 'Laptop',
            'High-performance business laptop', 'Dell', 'Latitude 5540', 'SN-DELL-001',
            'IN_USE', 'GOOD',
            locations[0] if locations else 'Head Office',
            departments[0] if departments else 'IT Department',
            'john.doe',
            vendors[0] if vendors else 'Tech Supplies Ltd',
            '2024-01-15', '2500.00', 'PO-2024-001', 'INV-2024-001', '2024-01-20',
            '2024-01-15', '2027-01-15', '36', '', '', '',
            '20', '5', '100.00',
            'Primary workstation', 'N', 'Y', 'POL-001', '2025-12-31',
        ],
        [
            'AST-002', 'Ergonomic Office Chair',
            categories[1] if len(categories) > 1 else (categories[0] if categories else 'Furniture'),
            asset_types[1] if len(asset_types) > 1 else (asset_types[0] if asset_types else 'Chair'),
            'Adjustable lumbar support chair', 'Steelcase', 'Leap V2', '',
            'IN_USE', 'EXCELLENT',
            locations[0] if locations else 'Head Office',
            departments[1] if len(departments) > 1 else (departments[0] if departments else 'Admin'),
            '',
            vendors[0] if vendors else 'Office Supplies PNG',
            '2024-03-10', '650.00', 'PO-2024-002', 'INV-2024-002', '2024-03-12',
            '', '2026-03-10', '24', '', '', '',
            '15', '8', '50.00',
            '', 'N', 'N', '', '',
        ],
    ]

    example_fill = hfill("FAFAFA")
    for row_idx, row_data in enumerate(example_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = Font(size=9, italic=True, color="666666")
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row_idx].height = 16

    # Row 6 – blank data start marker
    marker = ws.cell(row=6, column=1, value="↓ Enter your data below this row ↓")
    marker.font = Font(bold=True, color="C17845", size=9)
    marker.fill = hfill("FFF8F0")
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(COLS))
    ws.row_dimensions[6].height = 16

    # Style data rows 7–506 lightly
    data_fill = hfill("FFFFFF")
    for row_idx in range(7, 507):
        for col_idx in range(1, len(COLS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = data_fill
            ws.cell(row=row_idx, column=col_idx).border = thin_border
        ws.row_dimensions[row_idx].height = 15

    # ── Data Validation dropdowns ─────────────────────────────────────
    MAX_DATA_ROW = 506

    def add_validation(ws, formula, col_idx, first_row=7, last_row=MAX_DATA_ROW):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
        dv.error = "Value not in list. Check Reference sheet."
        dv.errorTitle = "Invalid value"
        col_letter = get_column_letter(col_idx)
        dv.sqref = f"{col_letter}{first_row}:{col_letter}{last_row}"
        ws.add_data_validation(dv)

    # Status (col 9) and Condition (col 10) – inline list
    status_formula = '"' + ','.join(statuses) + '"'
    condition_formula = '"' + ','.join(conditions) + '"'
    add_validation(ws, status_formula, 9)
    add_validation(ws, condition_formula, 10)

    # Is Critical / Is Insured (cols 30, 31)
    yn_formula = '"Y,N"'
    add_validation(ws, yn_formula, 30)
    add_validation(ws, yn_formula, 31)

    # ════════════════════════════════════════════════════════════════
    # SHEET 2 – Reference Data
    # ════════════════════════════════════════════════════════════════
    ws_ref = wb.create_sheet("Reference Data")
    ws_ref.freeze_panes = "A2"

    ref_sections = [
        ("Categories",  categories,  "A", "3F51B5"),
        ("Asset Types", asset_types, "C", "00897B"),
        ("Locations",   locations,   "E", "E65100"),
        ("Departments", departments, "G", "6A1B9A"),
        ("Vendors",     vendors,     "I", "AD1457"),
        ("Valid Statuses",   statuses,   "K", "37474F"),
        ("Valid Conditions", conditions, "M", "37474F"),
    ]

    for title, items, col_letter, color in ref_sections:
        col_idx = ord(col_letter) - ord('A') + 1
        hdr = ws_ref.cell(row=1, column=col_idx, value=title)
        hdr.fill = hfill(color)
        hdr.font = Font(bold=True, color=HEADER_FG, size=10)
        hdr.alignment = Alignment(horizontal='center', vertical='center')
        hdr.border = thin_border
        ws_ref.column_dimensions[col_letter].width = 24
        for row_idx, item in enumerate(items, start=2):
            cell = ws_ref.cell(row=row_idx, column=col_idx, value=item)
            cell.font = Font(size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    ws_ref.row_dimensions[1].height = 22

    # ── Named ranges for dropdown validation from Reference sheet ─────
    # (Only if company data exists – avoids empty range errors)
    def ref_range(col_letter, count):
        if count == 0:
            return None
        return f"'Reference Data'!${col_letter}$2:${col_letter}${count + 1}"

    ref_formulas = {
        4:  ref_range("A", len(categories)),   # Category col
        5:  ref_range("C", len(asset_types)),  # Asset Type col
        11: ref_range("E", len(locations)),    # Location col
        12: ref_range("G", len(departments)),  # Department col
        14: ref_range("I", len(vendors)),      # Vendor col
    }
    for col_idx, formula in ref_formulas.items():
        if formula:
            add_validation(ws, formula, col_idx)

    # ════════════════════════════════════════════════════════════════
    # SHEET 3 – Instructions
    # ════════════════════════════════════════════════════════════════
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.column_dimensions['A'].width = 30
    ws_inst.column_dimensions['B'].width = 65

    inst_title = ws_inst.cell(row=1, column=1, value="Asset Bulk Import – Instructions")
    inst_title.font = Font(bold=True, size=14, color=BRAND_DARK)
    ws_inst.merge_cells('A1:B1')
    ws_inst.row_dimensions[1].height = 24

    instructions = [
        ("", ""),
        ("HOW TO USE THIS TEMPLATE", ""),
        ("1. Go to 'Import Template' sheet", "Fill in your asset data starting from row 7 (rows 4-5 are examples – delete them before uploading)"),
        ("2. Check the 'Reference Data' sheet", "Use exact names from the Reference Data sheet for Category, Asset Type, Location, Department and Vendor columns"),
        ("3. Use dropdowns where available", "Status, Condition, Category, Asset Type, Location, Department and Vendor columns have dropdown lists"),
        ("4. Save as .xlsx", "Save the file in Excel format (.xlsx or .xls) before uploading"),
        ("5. Upload the file", "Go back to the Import page and upload your completed file"),
        ("", ""),
        ("REQUIRED FIELDS", ""),
        ("Asset Tag *",   "Unique identifier per company (e.g. AST-001). Duplicates will be skipped."),
        ("Asset Name *",  "Full descriptive name of the asset"),
        ("Category *",    "Must exactly match a category from the Reference Data sheet"),
        ("Asset Type *",  "Must exactly match an asset type from the Reference Data sheet"),
        ("", ""),
        ("DATE FORMAT", ""),
        ("Accepted formats",  "YYYY-MM-DD  (e.g. 2024-06-15)  OR  DD/MM/YYYY  (e.g. 15/06/2024)"),
        ("", ""),
        ("FINANCIAL FIELDS", ""),
        ("Purchase Price / AMC Cost / Salvage Value", "Enter numbers only, no currency symbols (e.g. 1500.00)"),
        ("Depreciation Rate %", "Enter the annual percentage as a number (e.g. 20 for 20%)"),
        ("Useful Life Years", "Integer number of years (e.g. 5)"),
        ("", ""),
        ("BOOLEAN FIELDS", ""),
        ("Is Critical / Is Insured", "Enter Y for Yes, N for No"),
        ("", ""),
        ("IMPORTANT NOTES", ""),
        ("Serial numbers", "Must be globally unique. Duplicates will be left blank with a warning."),
        ("Duplicate asset tags", "Rows with an existing Asset Tag will be skipped entirely."),
        ("Unknown references", "If a Category / Type / Location / Department name is not found, the system will use the first available value or skip the row."),
        ("Max rows", "Up to 500 assets can be imported per file."),
        ("File size", "Maximum file size: 10 MB"),
    ]

    for row_idx, (field, detail) in enumerate(instructions, start=2):
        c1 = ws_inst.cell(row=row_idx + 1, column=1, value=field)
        c2 = ws_inst.cell(row=row_idx + 1, column=2, value=detail)
        c2.alignment = Alignment(wrap_text=True, vertical='top')
        if field in ("HOW TO USE THIS TEMPLATE", "REQUIRED FIELDS", "DATE FORMAT", "FINANCIAL FIELDS", "BOOLEAN FIELDS", "IMPORTANT NOTES"):
            c1.font = Font(bold=True, size=11, color=BRAND_DARK)
            c1.fill = hfill("FFF3E0")
            c2.fill = hfill("FFF3E0")
        elif field.startswith(("1.", "2.", "3.", "4.", "5.")):
            c1.font = Font(bold=True, size=9)
        elif field.endswith("*"):
            c1.font = Font(bold=True, size=9, color="B71C1C")
        else:
            c1.font = Font(size=9)
        ws_inst.row_dimensions[row_idx + 1].height = 18

    # ── Save ──────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
