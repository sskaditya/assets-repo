from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
import json
import logging

logger = logging.getLogger(__name__)

from .models import Asset, AssetCategory, AssetType, Vendor, AssetDocument, AssetHistory
from .forms import (
    AssetForm, AssetFilterForm, AssetDocumentForm, AssetCategoryForm, 
    AssetTypeForm, VendorForm, AssetExcelImportForm
)
from users.models import Department, Location
from .utils import (
    calculate_current_book_value, 
    calculate_depreciation_schedule,
    get_assets_due_for_maintenance,
    get_assets_warranty_expiring,
    get_assets_amc_expiring,
    generate_asset_report_data
)


def _can_access_financial(request):
    """Finance Officers, Company Admins, and Super Admins can access financial views."""
    return (
        getattr(request, 'is_super_admin', False) or
        getattr(request, 'is_company_admin', False) or
        getattr(request, 'is_finance_officer', False)
    )


def _can_access_purchase(request):
    """Purchase Officers, Company Admins, and Super Admins can access vendor/purchase views."""
    return (
        getattr(request, 'is_super_admin', False) or
        getattr(request, 'is_company_admin', False) or
        getattr(request, 'is_purchase_officer', False)
    )


@login_required
def dashboard(request):
    """Dashboard view with asset statistics"""
    from datetime import date, timedelta
    from core.models import Company
    
    context = {}
    
    # Super Admin Dashboard - Company Management Focus
    if request.is_super_admin:
        # Company Statistics
        total_companies = Company.objects.filter(is_deleted=False).count()
        active_companies = Company.objects.filter(is_deleted=False, is_active=True).count()
        
        # Companies with expiring subscriptions (within 30 days)
        expiring_soon = Company.objects.filter(
            is_deleted=False,
            is_active=True,
            subscription_end_date__gte=date.today(),
            subscription_end_date__lte=date.today() + timedelta(days=30)
        ).count()
        
        # Recent companies
        recent_companies = Company.objects.filter(is_deleted=False).order_by('-created_at')[:5]
        
        # All assets across companies
        total_assets = Asset.objects.filter(is_deleted=False).count()
        
        # Companies list with asset counts
        companies_with_stats = Company.objects.filter(is_deleted=False).annotate(
            asset_count=Count('assets', filter=Q(assets__is_deleted=False)),
            user_count=Count('user_profiles', filter=Q(user_profiles__user__is_active=True))
        ).order_by('-created_at')[:10]

        # Assets belonging to expiring/expired companies
        expiring_company_ids = list(Company.objects.filter(
            is_deleted=False, is_active=True,
            subscription_end_date__gte=date.today(),
            subscription_end_date__lte=date.today() + timedelta(days=30),
        ).values_list('id', flat=True))
        expired_company_ids = list(Company.objects.filter(
            is_deleted=False, is_active=True,
            subscription_end_date__lt=date.today(),
        ).values_list('id', flat=True))
        assets_at_risk = Asset.objects.filter(
            company_id__in=expiring_company_ids + expired_company_ids,
            is_deleted=False,
        ).count()
        expired_companies_count = len(expired_company_ids)

        context.update({
            'is_super_admin_dashboard': True,
            'total_companies': total_companies,
            'active_companies': active_companies,
            'expiring_soon': expiring_soon,
            'recent_companies': recent_companies,
            'total_assets_all': total_assets,
            'companies_with_stats': companies_with_stats,
            'assets_at_risk': assets_at_risk,
            'expired_companies_count': expired_companies_count,
        })
    else:
        # Regular User Dashboard - Asset Management Focus
        company = getattr(request, 'current_company', None)
        
        if company:
            # Company-specific dashboard
            total_assets = Asset.objects.filter(company=company, is_deleted=False).count()
            active_assets = Asset.objects.filter(company=company, is_deleted=False, status='IN_USE').count()
            under_maintenance = Asset.objects.filter(company=company, is_deleted=False, status='UNDER_MAINTENANCE').count()
            critical_assets = Asset.objects.filter(company=company, is_deleted=False, is_critical=True).count()
            
            # Assets by category
            assets_by_category = AssetCategory.objects.filter(
                company=company, is_deleted=False, is_active=True
            ).annotate(
                asset_count=Count('assets', filter=Q(assets__is_deleted=False))
            ).order_by('-asset_count')[:5]
            
            # Recent assets
            recent_assets = Asset.objects.filter(company=company, is_deleted=False).select_related('category', 'location').order_by('-created_at')[:10]
            
            # Assets expiring warranty soon (within 30 days)
            warranty_expiring = Asset.objects.filter(
                company=company,
                is_deleted=False,
                warranty_end_date__gte=date.today(),
                warranty_end_date__lte=date.today() + timedelta(days=30)
            ).select_related('category', 'location').order_by('warranty_end_date')[:10]
        else:
            # No company context (shouldn't normally happen for regular users)
            total_assets = Asset.objects.filter(is_deleted=False).count()
            active_assets = Asset.objects.filter(is_deleted=False, status='IN_USE').count()
            under_maintenance = Asset.objects.filter(is_deleted=False, status='UNDER_MAINTENANCE').count()
            critical_assets = Asset.objects.filter(is_deleted=False, is_critical=True).count()
            
            # Assets by category
            assets_by_category = AssetCategory.objects.filter(
                is_deleted=False, is_active=True
            ).select_related('company').annotate(
                asset_count=Count('assets', filter=Q(assets__is_deleted=False))
            ).order_by('-asset_count')[:5]
            
            # Recent assets
            recent_assets = Asset.objects.filter(is_deleted=False).select_related('company', 'category', 'location').order_by('-created_at')[:10]
            
            # Assets expiring warranty soon (within 30 days)
            warranty_expiring = Asset.objects.filter(
                is_deleted=False,
                warranty_end_date__gte=date.today(),
                warranty_end_date__lte=date.today() + timedelta(days=30)
            ).select_related('company', 'category', 'location').order_by('warranty_end_date')[:10]
        
        # Portfolio depreciation summary
        depreciable_assets = Asset.objects.filter(
            **(dict(company=company) if company else dict()),
            is_deleted=False,
            purchase_price__isnull=False,
            purchase_date__isnull=False,
        )
        total_purchase_value = sum(a.purchase_price for a in depreciable_assets if a.purchase_price)
        total_current_value = sum(calculate_current_book_value(a) for a in depreciable_assets)
        total_depreciation_value = total_purchase_value - total_current_value
        depreciation_pct_overall = round(float(total_depreciation_value / total_purchase_value) * 100, 1) if total_purchase_value else 0
        retained_pct_overall = round(100 - depreciation_pct_overall, 1)

        # Top depreciating assets (highest % depreciated)
        top_depreciating = []
        for a in depreciable_assets:
            if a.purchase_price:
                cbv = calculate_current_book_value(a)
                dep_amt = a.purchase_price - cbv
                dep_pct = round(float(dep_amt / a.purchase_price) * 100, 1)
                top_depreciating.append({
                    'asset': a,
                    'current_value': cbv,
                    'depreciation_amount': dep_amt,
                    'depreciation_pct': dep_pct,
                })
        top_depreciating.sort(key=lambda x: x['depreciation_pct'], reverse=True)

        context.update({
            'is_super_admin_dashboard': False,
            'total_assets': total_assets,
            'active_assets': active_assets,
            'under_maintenance': under_maintenance,
            'critical_assets': critical_assets,
            'assets_by_category': assets_by_category,
            'recent_assets': recent_assets,
            'warranty_expiring': warranty_expiring,
            'company': company,
            'showing_all_companies': not company,
            'total_purchase_value': total_purchase_value,
            'total_current_value': total_current_value,
            'total_depreciation_value': total_depreciation_value,
            'depreciation_pct_overall': depreciation_pct_overall,
            'retained_pct_overall': retained_pct_overall,
            'top_depreciating': top_depreciating[:5],
        })
    
    return render(request, 'assets/dashboard.html', context)


@login_required
def asset_list(request):
    """List all assets with filtering"""
    company = getattr(request, 'current_company', None)
    
    # Filter by company if context exists
    if company:
        assets = Asset.objects.filter(company=company, is_deleted=False).select_related(
            'category', 'asset_type', 'location', 'department', 'assigned_to'
        ).order_by('-created_at')
    else:
        # Super admin - show all assets with company info
        assets = Asset.objects.filter(is_deleted=False).select_related(
            'company', 'category', 'asset_type', 'location', 'department', 'assigned_to'
        ).order_by('-created_at')
    
    # Apply filters
    filter_form = AssetFilterForm(request.GET)
    
    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        if search:
            assets = assets.filter(
                Q(asset_tag__icontains=search) |
                Q(name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(description__icontains=search)
            )
        
        category = filter_form.cleaned_data.get('category')
        if category:
            assets = assets.filter(category=category)
        
        asset_type = filter_form.cleaned_data.get('asset_type')
        if asset_type:
            assets = assets.filter(asset_type=asset_type)
        
        status = filter_form.cleaned_data.get('status')
        if status:
            assets = assets.filter(status=status)
        
        location = filter_form.cleaned_data.get('location')
        if location:
            assets = assets.filter(location=location)
        
        department = filter_form.cleaned_data.get('department')
        if department:
            assets = assets.filter(department=department)
        
        assigned_to = filter_form.cleaned_data.get('assigned_to')
        if assigned_to:
            assets = assets.filter(assigned_to=assigned_to)
    
    # Pagination
    paginator = Paginator(assets, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate depreciation for current page assets
    asset_depreciation_data = {}
    for asset in page_obj.object_list:
        if asset.purchase_price and asset.purchase_date:
            cbv = calculate_current_book_value(asset)
            dep_amt = asset.purchase_price - cbv
            dep_pct = round(float(dep_amt / asset.purchase_price) * 100, 1) if asset.purchase_price else 0
            asset_depreciation_data[asset.pk] = {
                'current_book_value': cbv,
                'depreciation_amount': dep_amt,
                'depreciation_pct': dep_pct,
            }

    context = {
        'page_obj': page_obj,
        'filter_form': filter_form,
        'company': company,
        'showing_all_companies': not company,
        'asset_depreciation_data': asset_depreciation_data,
    }

    return render(request, 'assets/asset_list.html', context)


@login_required
def asset_detail(request, pk):
    """Asset detail view"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)

    # Depreciation calculations
    current_book_value = None
    total_depreciation = None
    depreciation_pct = None
    depreciation_schedule = []
    years_elapsed = None

    if asset.purchase_price and asset.purchase_date:
        from datetime import date as _date
        current_book_value = calculate_current_book_value(asset)
        total_depreciation = asset.purchase_price - current_book_value
        depreciation_pct = round(float(total_depreciation / asset.purchase_price) * 100, 1) if asset.purchase_price else 0
        years_elapsed = round((_date.today() - asset.purchase_date).days / 365.25, 1)

    if asset.purchase_price and asset.purchase_date and asset.useful_life_years:
        depreciation_schedule = calculate_depreciation_schedule(asset)

    # Get related data
    documents = asset.documents.filter(is_deleted=False).order_by('-created_at')
    history = asset.history.all().order_by('-action_date')[:20]
    maintenance_logs = asset.maintenance_logs.filter(is_deleted=False).order_by('-maintenance_date')[:10]
    maintenance_requests = asset.maintenance_requests.filter(is_deleted=False).order_by('-requested_date')[:10]

    context = {
        'asset': asset,
        'documents': documents,
        'history': history,
        'maintenance_logs': maintenance_logs,
        'maintenance_requests': maintenance_requests,
        'current_book_value': current_book_value,
        'total_depreciation': total_depreciation,
        'depreciation_pct': depreciation_pct,
        'depreciation_schedule': depreciation_schedule,
        'years_elapsed': years_elapsed,
    }

    return render(request, 'assets/asset_detail.html', context)


@login_required
def asset_detail_by_qr(request, qr_code):
    """Asset detail view by QR code"""
    asset = get_object_or_404(Asset, qr_code=qr_code, is_deleted=False)
    return redirect('assets:asset_detail', pk=asset.pk)


@login_required
def asset_create(request):
    """Create new asset"""
    company = getattr(request, 'current_company', None)
    
    if not company:
        messages.error(request, 'Please select a specific company from the Company View selector before creating an asset.')
        return redirect('assets:asset_list')
    
    if request.method == 'POST':
        form = AssetForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                asset = form.save(commit=False)
                asset.company = company
                asset.save()
                
                AssetHistory.objects.create(
                    asset=asset,
                    action_type='CREATED',
                    performed_by=request.user,
                    remarks=f'Asset {asset.asset_tag} created'
                )
                
                messages.success(request, f'Asset {asset.asset_tag} created successfully!')
                return redirect('assets:asset_detail', pk=asset.pk)
            except IntegrityError:
                form.add_error(None, 'An asset with this tag or serial number already exists in this company.')
            except Exception as e:
                logger.error(f"Error creating asset: {e}")
                form.add_error(None, 'An unexpected error occurred while saving the asset. Please try again.')
    else:
        form = AssetForm()
    
    context = {
        'form': form,
        'title': 'Create New Asset',
    }
    
    return render(request, 'assets/asset_form.html', context)


@login_required
def asset_import_excel(request):
    """Import assets from Excel file"""
    from .excel_import import AssetExcelImporter, generate_import_template
    from django.http import HttpResponse
    
    company = getattr(request, 'current_company', None)
    
    if not company:
        messages.error(request, 'Company context is required for importing assets.')
        return redirect('assets:asset_list')
    
    # Handle template download
    if request.GET.get('download_template'):
        output = generate_import_template(company=company)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="asset_import_template.xlsx"'
        return response
    
    # Handle import
    if request.method == 'POST':
        form = AssetExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # Process import
            importer = AssetExcelImporter(excel_file, company, request.user)
            result = importer.import_assets()
            
            # Display results
            if result['success']:
                messages.success(
                    request,
                    f"Successfully imported {result['success_count']} asset(s). "
                    f"Skipped: {result['skip_count']}"
                )
            
            if result['errors']:
                for error in result['errors'][:5]:  # Show first 5 errors
                    messages.error(request, error)
                if len(result['errors']) > 5:
                    messages.warning(request, f"...and {len(result['errors']) - 5} more errors")
            
            if result['warnings']:
                for warning in result['warnings'][:3]:  # Show first 3 warnings
                    messages.warning(request, warning)
                if len(result['warnings']) > 3:
                    messages.info(request, f"...and {len(result['warnings']) - 3} more warnings")
            
            if result['success_count'] > 0:
                return redirect('assets:asset_list')
    else:
        form = AssetExcelImportForm()
    
    # Get statistics for display
    categories_count = AssetCategory.objects.filter(company=company, is_deleted=False).count()
    types_count = AssetType.objects.filter(company=company, is_deleted=False).count()
    locations_count = Location.objects.filter(company=company, is_deleted=False).count()
    
    context = {
        'form': form,
        'title': 'Import Assets from Excel',
        'categories_count': categories_count,
        'types_count': types_count,
        'locations_count': locations_count,
        'required_cols': [
            ('Asset Tag *',   'Unique identifier per company, e.g. AST-001'),
            ('Asset Name *',  'Full descriptive name of the asset'),
            ('Category *',    'Must match an existing category exactly'),
            ('Asset Type *',  'Must match an existing asset type exactly'),
        ],
        'basic_cols': [
            ('Description',           'Detailed description'),
            ('Make / Manufacturer',   'e.g. Dell, HP, Samsung'),
            ('Model',                 'e.g. Latitude 5540'),
            ('Serial Number',         'Must be globally unique'),
            ('Status',                'PLANNING · IN_USE · DEPLOYED · RETIRED …'),
            ('Condition',             'EXCELLENT · GOOD · FAIR · POOR · NOT_WORKING'),
            ('Location',              'Must match an existing location'),
            ('Department',            'Must match an existing department'),
            ('Assigned To',           'Username or email of the user'),
            ('Notes',                 'Any additional notes'),
            ('Is Critical (Y/N)',     'Y or N'),
            ('Is Insured (Y/N)',      'Y or N'),
            ('Insurance Policy No',   'Policy reference number'),
            ('Insurance Expiry',      'DD/MM/YYYY or YYYY-MM-DD'),
        ],
        'financial_cols': [
            ('Vendor',                'Must match an existing vendor'),
            ('Purchase Date',         'DD/MM/YYYY or YYYY-MM-DD'),
            ('Purchase Price',        'Numeric, e.g. 1500.00'),
            ('Purchase Order No',     'PO reference number'),
            ('Invoice Number',        'Invoice reference'),
            ('Invoice Date',          'DD/MM/YYYY or YYYY-MM-DD'),
            ('Warranty Start Date',   'DD/MM/YYYY or YYYY-MM-DD'),
            ('Warranty End Date',     'DD/MM/YYYY or YYYY-MM-DD'),
            ('Warranty Months',       'Number, e.g. 24'),
            ('AMC Start Date',        'DD/MM/YYYY or YYYY-MM-DD'),
            ('AMC End Date',          'DD/MM/YYYY or YYYY-MM-DD'),
            ('AMC Cost',              'Numeric, e.g. 500.00'),
        ],
        'depreciation_cols': [
            ('Depreciation Rate %', 'Annual rate, e.g. 20 for 20%'),
            ('Useful Life Years',   'Integer, e.g. 5'),
            ('Salvage Value',       'Numeric residual value, e.g. 100.00'),
        ],
    }
    
    return render(request, 'assets/asset_import_excel.html', context)


@login_required
def asset_update(request, pk):
    """Update asset"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        form = AssetForm(request.POST, request.FILES, instance=asset)
        if form.is_valid():
            try:
                old_status = asset.status
                old_location = asset.location
                old_assigned_to = asset.assigned_to
                
                asset = form.save()
                
                if old_status != asset.status:
                    AssetHistory.objects.create(
                        asset=asset,
                        action_type='STATUS_CHANGED',
                        performed_by=request.user,
                        old_value=old_status,
                        new_value=asset.status,
                        remarks=f'Status changed from {old_status} to {asset.status}'
                    )
                
                if old_location != asset.location:
                    AssetHistory.objects.create(
                        asset=asset,
                        action_type='LOCATION_CHANGED',
                        performed_by=request.user,
                        from_location=old_location,
                        to_location=asset.location,
                        remarks=f'Location changed'
                    )
                
                if old_assigned_to != asset.assigned_to:
                    AssetHistory.objects.create(
                        asset=asset,
                        action_type='ASSIGNED',
                        performed_by=request.user,
                        from_user=old_assigned_to,
                        to_user=asset.assigned_to,
                        remarks=f'Asset assigned to {asset.assigned_to.get_full_name() if asset.assigned_to else "Unassigned"}'
                    )
                
                messages.success(request, f'Asset {asset.asset_tag} updated successfully!')
                return redirect('assets:asset_detail', pk=asset.pk)
            except IntegrityError:
                form.add_error(None, 'An asset with this tag or serial number already exists in this company.')
            except Exception as e:
                logger.error(f"Error updating asset {asset.pk}: {e}")
                form.add_error(None, 'An unexpected error occurred while saving the asset. Please try again.')
    else:
        form = AssetForm(instance=asset)
    
    context = {
        'form': form,
        'asset': asset,
        'title': f'Update Asset: {asset.asset_tag}',
    }
    
    return render(request, 'assets/asset_form.html', context)


@login_required
def asset_delete(request, pk):
    """Soft delete asset"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        asset.soft_delete()
        
        # Create history entry
        AssetHistory.objects.create(
            asset=asset,
            action_type='DISPOSED',
            performed_by=request.user,
            remarks=f'Asset {asset.asset_tag} deleted'
        )
        
        messages.success(request, f'Asset {asset.asset_tag} deleted successfully!')
        return redirect('assets:asset_list')
    
    context = {
        'asset': asset,
    }
    
    return render(request, 'assets/asset_confirm_delete.html', context)


@login_required
def asset_transfer_create(request, pk):
    """Create an asset transfer request"""
    from .forms import AssetTransferForm
    from datetime import datetime
    
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    company = getattr(request, 'current_company', None)
    
    # Ensure user has access to this asset's company
    if company and asset.company != company:
        messages.error(request, 'You do not have permission to transfer this asset.')
        return redirect('assets:asset_detail', pk=pk)
    
    if request.method == 'POST':
        form = AssetTransferForm(request.POST, company=company, asset=asset)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.asset = asset
            transfer.requested_by = request.user
            transfer.from_user = asset.assigned_to
            transfer.from_location = asset.location
            transfer.from_department = asset.department
            transfer.status = 'PENDING'
            
            # Generate transfer number
            today = datetime.now()
            prefix = f"TR-{today.strftime('%Y%m')}"
            last_transfer = AssetTransfer.objects.filter(
                transfer_number__startswith=prefix
            ).order_by('-transfer_number').first()
            
            if last_transfer:
                last_number = int(last_transfer.transfer_number.split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1
            
            transfer.transfer_number = f"{prefix}-{new_number:04d}"
            transfer.save()
            
            # Create history entry
            AssetHistory.objects.create(
                asset=asset,
                action_type='TRANSFER_INITIATED',
                performed_by=request.user,
                remarks=f'Transfer request {transfer.transfer_number} created',
                from_user=transfer.from_user,
                to_user=transfer.to_user,
                from_location=transfer.from_location,
                to_location=transfer.to_location,
                from_department=transfer.from_department,
                to_department=transfer.to_department
            )
            
            messages.success(
                request,
                f'Transfer request {transfer.transfer_number} created successfully and is pending approval.'
            )
            return redirect('assets:asset_detail', pk=pk)
    else:
        form = AssetTransferForm(company=company, asset=asset)
    
    context = {
        'form': form,
        'asset': asset,
        'title': 'Transfer Asset',
    }
    
    return render(request, 'assets/asset_transfer_form.html', context)


@login_required
def asset_qr_code(request, pk):
    """View/download asset QR code"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    
    if not asset.qr_code_image:
        messages.error(request, 'QR code not available for this asset.')
        return redirect('assets:asset_detail', pk=asset.pk)
    
    # Return the QR code image
    from django.http import FileResponse
    return FileResponse(asset.qr_code_image.open(), content_type='image/png')


@login_required
def asset_add_document(request, pk):
    """Add document to asset"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        form = AssetDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.asset = asset
            document.uploaded_by = request.user
            document.save()
            
            messages.success(request, 'Document uploaded successfully!')
            return redirect('assets:asset_detail', pk=asset.pk)
    else:
        form = AssetDocumentForm()
    
    context = {
        'form': form,
        'asset': asset,
    }
    
    return render(request, 'assets/asset_document_form.html', context)


@login_required
def asset_categories(request):
    """List all asset categories – top-level with sub-categories nested"""
    company = getattr(request, 'current_company', None)

    base_qs = AssetCategory.objects.filter(is_deleted=False)
    if company:
        base_qs = base_qs.filter(company=company)
    else:
        base_qs = base_qs.select_related('company')

    # Top-level (parent) categories only
    top_level = base_qs.filter(parent_category__isnull=True).annotate(
        asset_count=Count('assets', filter=Q(assets__is_deleted=False)),
        sub_count=Count('sub_categories', filter=Q(sub_categories__is_deleted=False)),
    ).order_by('name')

    # Attach sub-categories to each parent (avoids N+1)
    all_subs = base_qs.filter(parent_category__isnull=False).annotate(
        asset_count=Count('assets', filter=Q(assets__is_deleted=False)),
    ).select_related('parent_category').order_by('name')

    sub_map = {}
    for sub in all_subs:
        sub_map.setdefault(sub.parent_category_id, []).append(sub)

    for cat in top_level:
        cat.children = sub_map.get(cat.pk, [])

    total_count     = base_qs.count()
    top_count       = top_level.count()
    sub_total_count = base_qs.filter(parent_category__isnull=False).count()

    context = {
        'categories': top_level,
        'company': company,
        'showing_all_companies': not company,
        'total_count': total_count,
        'top_count': top_count,
        'sub_total_count': sub_total_count,
    }

    return render(request, 'assets/category_list.html', context)


@login_required
def category_detail(request, pk):
    """Detail view for a single category, showing its sub-categories"""
    company = getattr(request, 'current_company', None)
    category = get_object_or_404(AssetCategory, pk=pk, is_deleted=False)

    sub_categories = AssetCategory.objects.filter(
        parent_category=category, is_deleted=False
    ).annotate(
        asset_count=Count('assets', filter=Q(assets__is_deleted=False))
    ).order_by('name')

    direct_assets = Asset.objects.filter(
        category=category, is_deleted=False
    ).select_related('asset_type', 'location')[:10]

    context = {
        'category': category,
        'sub_categories': sub_categories,
        'direct_assets': direct_assets,
        'company': company,
    }
    return render(request, 'assets/category_detail.html', context)


@login_required
def asset_types(request):
    """List all asset types"""
    company = getattr(request, 'current_company', None)
    
    if company:
        asset_types = AssetType.objects.filter(
            company=company, is_deleted=False, is_active=True
        ).select_related('category').annotate(
            asset_count=Count('assets', filter=Q(assets__is_deleted=False))
        ).order_by('category', 'name')
    else:
        # Super admin - show all types with company info
        asset_types = AssetType.objects.filter(
            is_deleted=False, is_active=True
        ).select_related('company', 'category').annotate(
            asset_count=Count('assets', filter=Q(assets__is_deleted=False))
        ).order_by('company__name', 'category', 'name')
    
    context = {
        'asset_types': asset_types,
        'company': company,
        'showing_all_companies': not company,
    }
    
    return render(request, 'assets/type_list.html', context)


@login_required
def vendors(request):
    """List all vendors"""
    if not _can_access_purchase(request):
        messages.error(request, 'You do not have permission to access vendor information.')
        return redirect('assets:dashboard')
    company = getattr(request, 'current_company', None)
    
    if company:
        vendors = Vendor.objects.filter(
            company=company, is_deleted=False, is_active=True
        ).order_by('name')
    else:
        # Super admin - show all vendors with company info
        vendors = Vendor.objects.filter(
            is_deleted=False, is_active=True
        ).select_related('company').order_by('company__name', 'name')
    
    context = {
        'vendors': vendors,
        'company': company,
        'showing_all_companies': not company,
    }
    
    return render(request, 'assets/vendor_list.html', context)


@login_required
def category_create(request):
    """Create new asset category"""
    company = getattr(request, 'current_company', None)
    
    if not company:
        messages.error(request, 'Please select a specific company from the Company View selector before creating a category.')
        return redirect('assets:category_list')
    
    if request.method == 'POST':
        form = AssetCategoryForm(request.POST, company=company)
        if form.is_valid():
            try:
                category = form.save(commit=False)
                category.company = company
                category.save()
                messages.success(request, f'Category "{category.name}" created successfully!')
                # Redirect back to parent detail page if we came from there
                parent = category.parent_category
                if parent:
                    return redirect('assets:category_detail', pk=parent.pk)
                return redirect('assets:category_list')
            except IntegrityError:
                form.add_error(None, 'A category with this code or name already exists for this company.')
            except Exception as e:
                logger.error(f"Error creating category: {e}")
                form.add_error(None, 'An unexpected error occurred. Please try again.')
    else:
        form = AssetCategoryForm(company=company)
        # Pre-select parent if ?parent=<pk> is passed (e.g. from "Add Sub-category" button)
        parent_pk = request.GET.get('parent')
        if parent_pk:
            try:
                parent_obj = AssetCategory.objects.get(pk=parent_pk, company=company, is_deleted=False)
                form.initial['parent_category'] = parent_obj
            except AssetCategory.DoesNotExist:
                pass

    context = {
        'form': form,
        'title': 'Create Asset Category',
    }
    
    return render(request, 'assets/category_form.html', context)


@login_required
def category_update(request, pk):
    """Update asset category"""
    category = get_object_or_404(AssetCategory, pk=pk, is_deleted=False)
    company = getattr(request, 'current_company', None)
    
    if request.method == 'POST':
        form = AssetCategoryForm(request.POST, instance=category, company=company)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" updated successfully!')
            return redirect('assets:category_list')
    else:
        form = AssetCategoryForm(instance=category, company=company)
    
    context = {
        'form': form,
        'category': category,
        'title': f'Update Category: {category.name}',
    }
    
    return render(request, 'assets/category_form.html', context)


@login_required
def category_delete(request, pk):
    """Delete asset category"""
    category = get_object_or_404(AssetCategory, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        category.soft_delete()
        messages.success(request, f'Category "{category.name}" deleted successfully!')
        return redirect('assets:category_list')
    
    context = {
        'category': category,
    }
    
    return render(request, 'assets/category_confirm_delete.html', context)


@login_required
def type_create(request):
    """Create new asset type"""
    company = getattr(request, 'current_company', None)
    
    if not company:
        messages.error(request, 'Please select a specific company from the Company View selector before creating an asset type.')
        return redirect('assets:type_list')
    
    if request.method == 'POST':
        form = AssetTypeForm(request.POST, company=company)
        if form.is_valid():
            try:
                asset_type = form.save(commit=False)
                asset_type.company = company
                asset_type.save()
                messages.success(request, f'Asset type "{asset_type.name}" created successfully!')
                return redirect('assets:type_list')
            except IntegrityError:
                form.add_error(None, 'An asset type with this code or name already exists for this company.')
            except Exception as e:
                logger.error(f"Error creating asset type: {e}")
                form.add_error(None, 'An unexpected error occurred. Please try again.')
    else:
        form = AssetTypeForm(company=company)
    
    context = {
        'form': form,
        'title': 'Create Asset Type',
    }
    
    return render(request, 'assets/type_form.html', context)


@login_required
def type_update(request, pk):
    """Update asset type"""
    asset_type = get_object_or_404(AssetType, pk=pk, is_deleted=False)
    company = getattr(request, 'current_company', None)
    
    if request.method == 'POST':
        form = AssetTypeForm(request.POST, instance=asset_type, company=company)
        if form.is_valid():
            asset_type = form.save()
            messages.success(request, f'Asset type "{asset_type.name}" updated successfully!')
            return redirect('assets:type_list')
    else:
        form = AssetTypeForm(instance=asset_type, company=company)
    
    context = {
        'form': form,
        'asset_type': asset_type,
        'title': f'Update Type: {asset_type.name}',
    }
    
    return render(request, 'assets/type_form.html', context)


@login_required
def type_delete(request, pk):
    """Delete asset type"""
    asset_type = get_object_or_404(AssetType, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        asset_type.soft_delete()
        messages.success(request, f'Asset type "{asset_type.name}" deleted successfully!')
        return redirect('assets:type_list')
    
    context = {
        'asset_type': asset_type,
    }
    
    return render(request, 'assets/type_confirm_delete.html', context)


@login_required
def vendor_create(request):
    """Create new vendor"""
    if not _can_access_purchase(request):
        messages.error(request, 'You do not have permission to create vendors.')
        return redirect('assets:vendor_list')
    company = getattr(request, 'current_company', None)

    if not company:
        messages.error(request, 'Please select a specific company from the Company View selector before creating a vendor.')
        return redirect('assets:vendor_list')
    
    if request.method == 'POST':
        form = VendorForm(request.POST)
        if form.is_valid():
            try:
                vendor = form.save(commit=False)
                vendor.company = company
                vendor.save()
                messages.success(request, f'Vendor "{vendor.name}" created successfully!')
                return redirect('assets:vendor_list')
            except IntegrityError:
                form.add_error(None, 'A vendor with this code or name already exists for this company.')
            except Exception as e:
                logger.error(f"Error creating vendor: {e}")
                form.add_error(None, 'An unexpected error occurred. Please try again.')
    else:
        form = VendorForm()
    
    context = {
        'form': form,
        'title': 'Create Vendor',
    }
    
    return render(request, 'assets/vendor_form.html', context)


@login_required
def vendor_update(request, pk):
    """Update vendor"""
    if not _can_access_purchase(request):
        messages.error(request, 'You do not have permission to update vendors.')
        return redirect('assets:vendor_list')
    vendor = get_object_or_404(Vendor, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            vendor = form.save()
            messages.success(request, f'Vendor "{vendor.name}" updated successfully!')
            return redirect('assets:vendor_list')
    else:
        form = VendorForm(instance=vendor)
    
    context = {
        'form': form,
        'vendor': vendor,
        'title': f'Update Vendor: {vendor.name}',
    }
    
    return render(request, 'assets/vendor_form.html', context)


@login_required
def vendor_delete(request, pk):
    """Delete vendor"""
    if not _can_access_purchase(request):
        messages.error(request, 'You do not have permission to delete vendors.')
        return redirect('assets:vendor_list')
    vendor = get_object_or_404(Vendor, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        vendor.soft_delete()
        messages.success(request, f'Vendor "{vendor.name}" deleted successfully!')
        return redirect('assets:vendor_list')
    
    context = {
        'vendor': vendor,
    }
    
    return render(request, 'assets/vendor_confirm_delete.html', context)


# ============================================
# QR Code Scanning & Asset Tracking
# ============================================

@login_required
def qr_scanner(request):
    """QR code scanner for mobile and handheld devices"""
    return render(request, 'assets/qr_scanner.html')


@login_required
def asset_lookup_api(request):
    """API endpoint for asset lookup by QR code or asset tag"""
    code = request.GET.get('code', '')
    company = getattr(request, 'current_company', None)
    
    if not code:
        return JsonResponse({'success': False, 'message': 'No code provided'})
    
    # Build base query with company filter if applicable
    base_query = {'is_deleted': False}
    if company:
        base_query['company'] = company
    
    # Try to find asset by QR code (UUID) or asset tag
    asset = None
    try:
        # Try UUID lookup first
        asset = Asset.objects.get(qr_code=code, **base_query)
    except:
        # Try asset tag lookup
        try:
            asset = Asset.objects.get(asset_tag=code, **base_query)
        except:
            pass
    
    if not asset:
        return JsonResponse({'success': False, 'message': 'Asset not found'})
    
    # Return asset data
    data = {
        'success': True,
        'asset': {
            'id': asset.id,
            'asset_tag': asset.asset_tag,
            'name': asset.name,
            'serial_number': asset.serial_number,
            'status': asset.get_status_display(),
            'condition': asset.get_condition_display() if asset.condition else None,
            'category': asset.category.name if asset.category else None,
            'location': asset.location.name if asset.location else None,
            'location_id': asset.location.id if asset.location else None,
            'assigned_to': asset.assigned_to.get_full_name() if asset.assigned_to else None,
        }
    }
    
    # Record scan in history
    AssetHistory.objects.create(
        asset=asset,
        action_type='UPDATED',
        performed_by=request.user,
        remarks=f'Asset scanned via QR code by {request.user.get_full_name()}'
    )
    
    return JsonResponse(data)


@login_required
def locations_api(request):
    """API endpoint to get all locations"""
    locations = Location.objects.filter(is_deleted=False, is_active=True).values('id', 'name', 'code')
    return JsonResponse({'locations': list(locations)})


@login_required
def update_asset_condition(request):
    """Update asset condition via API"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required'})
    
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')
        condition = data.get('condition')
        remarks = data.get('remarks', '')
        
        if not asset_id or not condition:
            return JsonResponse({'success': False, 'message': 'Asset ID and condition are required'})
        
        valid_conditions = [c[0] for c in Asset.CONDITION_CHOICES]
        if condition not in valid_conditions:
            return JsonResponse({'success': False, 'message': 'Invalid condition value'})
        
        asset = get_object_or_404(Asset, pk=asset_id, is_deleted=False)
        old_condition = asset.condition
        asset.condition = condition
        asset.save()
        
        AssetHistory.objects.create(
            asset=asset,
            action_type='UPDATED',
            performed_by=request.user,
            old_value=old_condition,
            new_value=condition,
            remarks=remarks or f'Condition updated to {asset.get_condition_display()} by {request.user.get_full_name()}'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Condition updated successfully',
            'condition': asset.get_condition_display()
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def record_asset_movement(request):
    """Record asset movement from one location to another"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required'})
    
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')
        to_location_id = data.get('to_location_id')
        remarks = data.get('remarks', '')
        
        asset = get_object_or_404(Asset, pk=asset_id, is_deleted=False)
        to_location = get_object_or_404(Location, pk=to_location_id, is_deleted=False)
        from_location = asset.location
        
        # Update asset location
        asset.location = to_location
        asset.save()
        
        # Record in history
        AssetHistory.objects.create(
            asset=asset,
            action_type='LOCATION_CHANGED',
            performed_by=request.user,
            from_location=from_location,
            to_location=to_location,
            remarks=remarks or f'Asset moved by {request.user.get_full_name()}'
        )
        
        return JsonResponse({'success': True, 'message': 'Movement recorded successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ============================================
# Financial Management & Depreciation
# ============================================

@login_required
def financial_dashboard(request):
    """Financial dashboard with depreciation calculations"""
    if not _can_access_financial(request):
        messages.error(request, 'You do not have permission to access financial information.')
        return redirect('assets:dashboard')
    company = getattr(request, 'current_company', None)
    
    if company:
        assets = Asset.objects.filter(
            company=company, is_deleted=False, purchase_price__isnull=False
        ).select_related('category', 'location')
    else:
        # Super admin - show all assets
        assets = Asset.objects.filter(
            is_deleted=False, purchase_price__isnull=False
        ).select_related('category', 'location', 'company')
    
    # Calculate totals
    total_purchase_value = sum(asset.purchase_price for asset in assets if asset.purchase_price)
    total_current_value = sum(calculate_current_book_value(asset) for asset in assets)
    total_depreciation = total_purchase_value - total_current_value
    
    # Calculate percentages
    if total_purchase_value > 0:
        depreciation_percentage = round((total_depreciation / total_purchase_value) * 100, 1)
        current_value_percentage = round((total_current_value / total_purchase_value) * 100, 1)
    else:
        depreciation_percentage = 0
        current_value_percentage = 0
    
    # Assets with high depreciation
    high_depreciation_assets = []
    for asset in assets:
        if asset.purchase_price:
            current_value = calculate_current_book_value(asset)
            depreciation_amount = asset.purchase_price - current_value
            depreciation_pct = ((asset.purchase_price - current_value) / asset.purchase_price) * 100
            if depreciation_pct > 50:
                high_depreciation_assets.append({
                    'asset': asset,
                    'depreciation_pct': round(depreciation_pct, 2),
                    'depreciation_amount': round(depreciation_amount, 2),
                    'current_value': round(current_value, 2)
                })
    
    # Sort by depreciation percentage
    high_depreciation_assets.sort(key=lambda x: x['depreciation_pct'], reverse=True)
    
    context = {
        'total_purchase_value': total_purchase_value,
        'total_current_value': total_current_value,
        'total_depreciation': total_depreciation,
        'depreciation_percentage': depreciation_percentage,
        'current_value_percentage': current_value_percentage,
        'high_depreciation_assets': high_depreciation_assets[:10],
        'company': company,
        'showing_all_companies': not company,
    }
    
    return render(request, 'assets/financial_dashboard.html', context)


@login_required
def asset_depreciation(request, pk):
    """View asset depreciation schedule"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    
    current_value = calculate_current_book_value(asset)
    schedule = calculate_depreciation_schedule(asset)
    
    context = {
        'asset': asset,
        'current_value': current_value,
        'schedule': schedule,
    }
    
    return render(request, 'assets/asset_depreciation.html', context)


# ============================================
# Asset Monitoring & Control
# ============================================

@login_required
def monitoring_dashboard(request):
    """Real-time asset monitoring dashboard with detailed breakdowns"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    from django.db.models import Q, Sum, Avg
    
    # Get current company
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Base queryset
    if company:
        base_assets = Asset.objects.filter(company=company, is_deleted=False)
    else:
        # Super admin or no company context - show all assets
        base_assets = Asset.objects.filter(is_deleted=False)
    
    # Company-wise summary (for super admin)
    if not company:
        company_summary = base_assets.values(
            'company__id',
            'company__name',
            'company__code'
        ).annotate(
            total_assets=Count('id'),
            planning=Count('id', filter=Q(status='PLANNING')),
            ordered=Count('id', filter=Q(status='ORDERED')),
            in_transit=Count('id', filter=Q(status='IN_TRANSIT')),
            received=Count('id', filter=Q(status='RECEIVED')),
            available=Count('id', filter=Q(status='AVAILABLE')),
            deployed=Count('id', filter=Q(status='DEPLOYED')),
            in_use=Count('id', filter=Q(status='IN_USE')),
            under_maintenance=Count('id', filter=Q(status='UNDER_MAINTENANCE')),
            retired=Count('id', filter=Q(status='RETIRED')),
            disposed=Count('id', filter=Q(status='DISPOSED')),
            lost=Count('id', filter=Q(status='LOST')),
            stolen=Count('id', filter=Q(status='STOLEN')),
            critical_assets=Count('id', filter=Q(is_critical=True)),
            total_value=Sum('purchase_price'),
            avg_value=Avg('purchase_price')
        ).order_by('-total_assets')
    else:
        company_summary = None
    
    # Asset Type breakdown
    asset_type_summary = base_assets.values(
        'asset_type__id',
        'asset_type__name',
        'asset_type__code',
        'category__name'
    ).annotate(
        total_count=Count('id'),
        planning=Count('id', filter=Q(status='PLANNING')),
        available=Count('id', filter=Q(status='AVAILABLE')),
        deployed=Count('id', filter=Q(status='DEPLOYED')),
        in_use=Count('id', filter=Q(status='IN_USE')),
        under_maintenance=Count('id', filter=Q(status='UNDER_MAINTENANCE')),
        critical=Count('id', filter=Q(is_critical=True)),
        total_value=Sum('purchase_price')
    ).order_by('-total_count')
    
    if company:
        # Add company info for single company view
        for item in asset_type_summary:
            item['company_name'] = company.name
    else:
        # For super admin, add company name to each type
        for item in asset_type_summary:
            # Get company from first asset of this type
            first_asset = base_assets.filter(asset_type__id=item['asset_type__id']).select_related('company').first()
            if first_asset:
                item['company_name'] = first_asset.company.name if first_asset.company else 'N/A'
    
    # Procurement Status (Planning, Ordered, In Transit)
    procurement_assets = base_assets.filter(
        status__in=['PLANNING', 'ORDERED', 'IN_TRANSIT', 'RECEIVED']
    ).select_related('company', 'asset_type', 'category', 'vendor')
    
    # Critical Assets Detail
    critical_assets_detail = base_assets.filter(
        is_critical=True
    ).select_related('company', 'asset_type', 'location', 'assigned_to').order_by('-purchase_price')
    
    # Status Summary with percentage calculation
    total_assets_count = base_assets.count()
    status_summary = base_assets.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Add percentage to status summary
    for item in status_summary:
        if total_assets_count > 0:
            item['percentage'] = round((item['count'] / total_assets_count) * 100, 1)
        else:
            item['percentage'] = 0
    
    # Assets under maintenance
    under_maintenance = base_assets.filter(
        status='UNDER_MAINTENANCE'
    ).select_related('location', 'company', 'asset_type')
    
    # Upcoming maintenance, warranty, AMC
    today = timezone.now().date()
    if company:
        maintenance_due = get_assets_due_for_maintenance(company, days_ahead=7)
        warranty_expiring = get_assets_warranty_expiring(company, days_ahead=30)
        amc_expiring = get_assets_amc_expiring(company, days_ahead=30)
    else:
        maintenance_due = base_assets.filter(
            warranty_end_date__gte=today,
            warranty_end_date__lte=today + timedelta(days=7)
        ).select_related('category', 'location', 'company')
        
        warranty_expiring = base_assets.filter(
            warranty_end_date__gte=today,
            warranty_end_date__lte=today + timedelta(days=30)
        ).select_related('category', 'location', 'company', 'asset_type')
        
        amc_expiring = base_assets.filter(
            amc_end_date__gte=today,
            amc_end_date__lte=today + timedelta(days=30)
        ).select_related('category', 'location', 'amc_vendor', 'company', 'asset_type')
    
    # Recent asset activities
    if company:
        recent_activities = AssetHistory.objects.filter(
            asset__company=company
        ).select_related('asset', 'asset__asset_type', 'performed_by').order_by('-action_date')[:20]
    else:
        recent_activities = AssetHistory.objects.select_related(
            'asset', 'asset__company', 'asset__asset_type', 'performed_by'
        ).order_by('-action_date')[:20]
    
    context = {
        'company_summary': company_summary,
        'asset_type_summary': asset_type_summary,
        'procurement_assets': procurement_assets,
        'critical_assets_detail': critical_assets_detail,
        'status_summary': status_summary,
        'under_maintenance': under_maintenance,
        'maintenance_due': maintenance_due,
        'warranty_expiring': warranty_expiring,
        'amc_expiring': amc_expiring,
        'recent_activities': recent_activities,
        'company': company,
        'showing_all_companies': not company,
        'is_super_admin': is_super_admin,
    }
    
    return render(request, 'assets/monitoring_dashboard.html', context)


# ============================================
# Complete Audit Trail System
# ============================================

@login_required
def audit_trail(request):
    """Complete audit trail with advanced filtering"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    from django.contrib.auth.models import User
    
    # Get current company
    company = getattr(request, 'current_company', None)
    
    # Base query
    if company:
        history = AssetHistory.objects.filter(
            asset__company=company
        ).select_related('asset', 'performed_by', 'from_location', 'to_location', 'from_user', 'to_user')
    else:
        # Super admin sees all
        history = AssetHistory.objects.all().select_related(
            'asset', 'asset__company', 'performed_by', 'from_location', 'to_location', 'from_user', 'to_user'
        )
    
    # Apply filters
    asset_id = request.GET.get('asset_id')
    if asset_id:
        history = history.filter(asset_id=asset_id)
    
    action_type = request.GET.get('action_type')
    if action_type:
        history = history.filter(action_type=action_type)
    
    user_id = request.GET.get('user_id')
    if user_id:
        history = history.filter(performed_by_id=user_id)
    
    # Date range filter
    date_range = request.GET.get('date_range', 'all')
    today = timezone.now().date()
    
    if date_range == 'today':
        history = history.filter(action_date__date=today)
    elif date_range == 'week':
        week_ago = today - timedelta(days=7)
        history = history.filter(action_date__date__gte=week_ago)
    elif date_range == 'month':
        month_ago = today - timedelta(days=30)
        history = history.filter(action_date__date__gte=month_ago)
    elif date_range == 'quarter':
        quarter_ago = today - timedelta(days=90)
        history = history.filter(action_date__date__gte=quarter_ago)
    
    # Custom date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        history = history.filter(action_date__date__gte=start_date)
    if end_date:
        history = history.filter(action_date__date__lte=end_date)
    
    # Order by latest first
    history = history.order_by('-action_date')
    
    # Statistics
    total_actions = history.count()
    assets_modified = history.values('asset').distinct().count()
    users_active = history.values('performed_by').distinct().count()
    today_actions = history.filter(action_date__date=today).count()
    
    # Get all assets and users for filters
    if company:
        all_assets = Asset.objects.filter(company=company, is_deleted=False).order_by('asset_tag')
        all_users = User.objects.filter(profile__company=company).order_by('first_name', 'last_name')
    else:
        all_assets = Asset.objects.filter(is_deleted=False).order_by('asset_tag')
        all_users = User.objects.all().order_by('first_name', 'last_name')
    
    # Pagination
    paginator = Paginator(history, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_actions': total_actions,
        'assets_modified': assets_modified,
        'users_active': users_active,
        'today_actions': today_actions,
        'all_assets': all_assets,
        'all_users': all_users,
    }
    
    return render(request, 'assets/audit_trail.html', context)


@login_required
def company_audit_report(request):
    """Company-wise audit report"""
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    from core.models import Company
    
    # Must be super admin
    if not getattr(request, 'is_super_admin', False):
        messages.error(request, 'Only super admin can access this report')
        return redirect('assets:audit_trail')
    
    # Get all companies
    companies = Company.objects.filter(is_deleted=False, is_active=True).annotate(
        total_actions=Count('assets__history'),
        actions_this_month=Count('assets__history', filter=Q(
            assets__history__action_date__gte=timezone.now() - timedelta(days=30)
        )),
        total_assets=Count('assets', filter=Q(assets__is_deleted=False))
    ).order_by('name')
    
    # Get date range
    date_range = request.GET.get('date_range', 'month')
    today = timezone.now().date()
    
    if date_range == 'week':
        start_date = today - timedelta(days=7)
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
    elif date_range == 'quarter':
        start_date = today - timedelta(days=90)
    elif date_range == 'year':
        start_date = today - timedelta(days=365)
    else:
        start_date = None
    
    # Get detailed company data
    company_data = []
    for company in companies:
        history_query = AssetHistory.objects.filter(asset__company=company)
        
        if start_date:
            history_query = history_query.filter(action_date__date__gte=start_date)
        
        actions_by_type = history_query.values('action_type').annotate(count=Count('id'))
        
        company_data.append({
            'company': company,
            'actions_by_type': actions_by_type,
            'most_active_asset': history_query.values('asset__asset_tag', 'asset__name').annotate(
                count=Count('id')
            ).order_by('-count').first(),
            'most_active_user': history_query.values(
                'performed_by__first_name', 'performed_by__last_name', 'performed_by__username'
            ).annotate(count=Count('id')).order_by('-count').first()
        })
    
    context = {
        'companies': companies,
        'company_data': company_data,
        'date_range': date_range,
    }
    
    return render(request, 'assets/company_audit_report.html', context)


@login_required
def asset_audit_detail(request, pk):
    """Detailed audit trail for a specific asset"""
    asset = get_object_or_404(Asset, pk=pk, is_deleted=False)
    
    # Get complete history
    history = asset.history.all().select_related(
        'performed_by', 'from_location', 'to_location', 'from_user', 'to_user'
    ).order_by('-action_date')
    
    # Statistics
    total_actions = history.count()
    location_changes = history.filter(action_type='LOCATION_CHANGED').count()
    assignments = history.filter(action_type='ASSIGNED').count()
    status_changes = history.filter(action_type='STATUS_CHANGED').count()
    maintenance_events = history.filter(action_type='MAINTENANCE').count()
    
    # Timeline data
    timeline = []
    for entry in history:
        timeline.append({
            'date': entry.action_date,
            'action': entry.get_action_type_display(),
            'user': entry.performed_by.get_full_name() if entry.performed_by else 'System',
            'details': entry.remarks or '',
            'icon': get_action_icon(entry.action_type)
        })
    
    context = {
        'asset': asset,
        'history': history,
        'total_actions': total_actions,
        'location_changes': location_changes,
        'assignments': assignments,
        'status_changes': status_changes,
        'maintenance_events': maintenance_events,
        'timeline': timeline,
    }
    
    return render(request, 'assets/asset_audit_detail.html', context)


@login_required
def audit_export(request):
    """Export audit trail to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.utils import timezone
    from datetime import datetime
    
    # Get filtered data (same as audit_trail view)
    company = getattr(request, 'current_company', None)
    
    if company:
        history = AssetHistory.objects.filter(
            asset__company=company
        ).select_related('asset', 'performed_by', 'from_location', 'to_location')
    else:
        history = AssetHistory.objects.all().select_related(
            'asset', 'asset__company', 'performed_by', 'from_location', 'to_location'
        )
    
    # Apply same filters as audit_trail view
    asset_id = request.GET.get('asset_id')
    if asset_id:
        history = history.filter(asset_id=asset_id)
    
    action_type = request.GET.get('action_type')
    if action_type:
        history = history.filter(action_type=action_type)
    
    user_id = request.GET.get('user_id')
    if user_id:
        history = history.filter(performed_by_id=user_id)
    
    start_date = request.GET.get('start_date')
    if start_date:
        history = history.filter(action_date__date__gte=start_date)
    
    end_date = request.GET.get('end_date')
    if end_date:
        history = history.filter(action_date__date__lte=end_date)
    
    history = history.order_by('-action_date')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    
    # Header style
    header_fill = PatternFill(start_color="C17845", end_color="C17845", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Date & Time', 'Asset Tag', 'Asset Name', 'Action Type', 'User', 'From', 'To', 'Remarks', 'Company']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row, entry in enumerate(history, 2):
        ws.cell(row=row, column=1, value=entry.action_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=entry.asset.asset_tag)
        ws.cell(row=row, column=3, value=entry.asset.name)
        ws.cell(row=row, column=4, value=entry.get_action_type_display())
        ws.cell(row=row, column=5, value=entry.performed_by.get_full_name() if entry.performed_by else 'System')
        ws.cell(row=row, column=6, value=entry.from_location.name if entry.from_location else '')
        ws.cell(row=row, column=7, value=entry.to_location.name if entry.to_location else '')
        ws.cell(row=row, column=8, value=entry.remarks or '')
        ws.cell(row=row, column=9, value=entry.asset.company.name if hasattr(entry.asset, 'company') else '')
    
    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=audit_trail_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


def get_action_icon(action_type):
    """Helper function to get icon for action type"""
    icons = {
        'CREATED': 'plus',
        'UPDATED': 'edit',
        'ASSIGNED': 'user-check',
        'TRANSFERRED': 'arrow-right',
        'RETURNED': 'arrow-left',
        'MAINTENANCE': 'wrench',
        'REPAIRED': 'check-circle',
        'STATUS_CHANGED': 'alert-circle',
        'LOCATION_CHANGED': 'map-pin',
        'DISPOSED': 'trash-2',
    }
    return icons.get(action_type, 'circle')


# ============================================
# Comprehensive Reporting & Analytics System
# ============================================

@login_required
def reports_dashboard(request):
    """Main reporting dashboard with all report types"""
    company = getattr(request, 'current_company', None)
    
    # Get available report categories
    report_categories = [
        {
            'name': 'Asset Reports',
            'icon': 'package',
            'reports': [
                {'title': 'Asset Summary', 'url': 'assets:report_asset_summary', 'description': 'Overview of all assets'},
                {'title': 'Asset List', 'url': 'assets:report_asset_list', 'description': 'Detailed asset listing'},
                {'title': 'Asset by Category', 'url': 'assets:report_by_category', 'description': 'Categorized breakdown'},
                {'title': 'Asset by Location', 'url': 'assets:report_by_location', 'description': 'Location-wise distribution'},
            ]
        },
        {
            'name': 'Financial Reports',
            'icon': 'dollar-sign',
            'reports': [
                {'title': 'Financial Summary', 'url': 'assets:report_financial', 'description': 'Asset values and depreciation'},
                {'title': 'Depreciation Schedule', 'url': 'assets:report_depreciation', 'description': 'Depreciation breakdown'},
                {'title': 'Disposal Report', 'url': 'assets:report_disposal', 'description': 'Asset disposal tracking'},
            ]
        },
        {
            'name': 'Maintenance Reports',
            'icon': 'wrench',
            'reports': [
                {'title': 'Warranty Report', 'url': 'assets:report_warranty', 'description': 'Warranty status and expiry'},
                {'title': 'AMC Report', 'url': 'assets:report_amc', 'description': 'AMC tracking and renewal'},
                {'title': 'Maintenance Schedule', 'url': 'assets:report_maintenance', 'description': 'Maintenance tracking'},
            ]
        },
        {
            'name': 'Operational Reports',
            'icon': 'activity',
            'reports': [
                {'title': 'Transfer Report', 'url': 'assets:report_transfer', 'description': 'Asset transfer tracking'},
                {'title': 'Assignment Report', 'url': 'assets:report_assignment', 'description': 'Asset assignments'},
                {'title': 'Movement Report', 'url': 'assets:report_movement', 'description': 'Asset movement history'},
            ]
        },
    ]
    
    context = {
        'report_categories': report_categories,
    }
    
    return render(request, 'assets/reports_dashboard.html', context)


@login_required
def report_asset_summary(request):
    """Asset Summary Report"""
    from .reports import AssetSummaryReport
    from django.db.models import Count, Sum, Avg, Q
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = AssetSummaryReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        
        company_data = []
        for comp in companies:
            assets = Asset.objects.filter(company=comp, is_deleted=False)
            if assets.exists():
                company_data.append({
                    'company': comp,
                    'total_assets': assets.count(),
                    'critical_assets': assets.filter(is_critical=True).count(),
                    'total_value': assets.aggregate(total=Sum('purchase_price'))['total'] or 0,
                    'avg_value': assets.aggregate(avg=Avg('purchase_price'))['avg'] or 0,
                    'by_status': assets.values('status').annotate(count=Count('id')).order_by('-count')[:5],
                    'by_condition': assets.values('condition').annotate(count=Count('id')).order_by('-count')[:3],
                })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = AssetSummaryReport(company=company)
        data = report.generate()
        
        context = {
            'report_data': data,
            'report_title': 'Asset Summary Report',
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_asset_summary.html', context)


@login_required
def report_asset_list(request):
    """Detailed Asset List Report"""
    from .reports import AssetListReport
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Get filter options
    filters = {
        'status': request.GET.get('status'),
        'category': request.GET.get('category'),
        'location': request.GET.get('location'),
        'department': request.GET.get('department'),
        'condition': request.GET.get('condition'),
    }
    
    if request.GET.get('export') == 'excel':
        report = AssetListReport(company=company)
        return report.export_to_excel(filters=filters)
    
    report = AssetListReport(company=company)
    assets = report.generate(filters=filters)
    
    # Pagination
    paginator = Paginator(assets, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter dropdowns
    if company:
        categories = AssetCategory.objects.filter(company=company, is_deleted=False)
        locations = Location.objects.filter(company=company, is_deleted=False)
        departments = Department.objects.filter(company=company, is_deleted=False)
    else:
        categories = AssetCategory.objects.filter(is_deleted=False)
        locations = Location.objects.filter(is_deleted=False)
        departments = Department.objects.filter(is_deleted=False)
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'locations': locations,
        'departments': departments,
        'filters': filters,
        'company': company,
        'is_super_admin': is_super_admin,
        'showing_all_companies': not company and is_super_admin,
    }
    
    return render(request, 'assets/report_asset_list.html', context)


@login_required
def report_financial(request):
    """Financial Report"""
    if not _can_access_financial(request):
        messages.error(request, 'You do not have permission to access financial reports.')
        return redirect('assets:dashboard')
    from .reports import FinancialReport

    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = FinancialReport(company=company)
        return report.export_to_excel()
    
    report = FinancialReport(company=company)
    data = report.generate()
    
    # Pagination
    paginator = Paginator(data['asset_financials'], 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_purchase_value': data['total_purchase_value'],
        'total_book_value': data['total_book_value'],
        'total_depreciation': data['total_depreciation'],
        'company': company,
        'is_super_admin': is_super_admin,
        'showing_all_companies': not company and is_super_admin,
    }
    
    return render(request, 'assets/report_financial.html', context)


@login_required
def report_maintenance(request):
    """Maintenance Report"""
    from .reports import MaintenanceReport
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = MaintenanceReport(company=company)
        return report.export_to_excel()
    
    report = MaintenanceReport(company=company)
    data = report.generate()
    
    context = {
        'warranty_expiring': data['warranty_expiring'],
        'warranty_expired': data['warranty_expired'][:10],  # Limit to 10
        'amc_expiring': data['amc_expiring'],
        'amc_expired': data['amc_expired'][:10],  # Limit to 10
        'under_warranty': data['under_warranty'],
        'under_amc': data['under_amc'],
        'company': company,
        'is_super_admin': is_super_admin,
        'showing_all_companies': not company and is_super_admin,
    }
    
    return render(request, 'assets/report_maintenance.html', context)


@login_required
def report_transfer(request):
    """Transfer Report"""
    from .reports import TransferReport
    from datetime import datetime, timedelta
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = datetime.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if request.GET.get('export') == 'excel':
        report = TransferReport(company=company, start_date=start_date, end_date=end_date)
        return report.export_to_excel()
    
    report = TransferReport(company=company, start_date=start_date, end_date=end_date)
    data = report.generate()
    
    # Pagination
    paginator = Paginator(data['all_transfers'], 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'pending_count': data['pending_count'],
        'approved_count': data['approved_count'],
        'completed_count': data['completed_count'],
        'rejected_count': data['rejected_count'],
        'start_date': start_date,
        'end_date': end_date,
        'company': company,
        'is_super_admin': is_super_admin,
        'showing_all_companies': not company and is_super_admin,
    }
    
    return render(request, 'assets/report_transfer.html', context)


@login_required
def report_disposal(request):
    """Disposal Report"""
    from .reports import DisposalReport
    from datetime import datetime, timedelta
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = datetime.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if request.GET.get('export') == 'excel':
        report = DisposalReport(company=company, start_date=start_date, end_date=end_date)
        return report.export_to_excel()
    
    report = DisposalReport(company=company, start_date=start_date, end_date=end_date)
    data = report.generate()
    
    # Pagination
    paginator = Paginator(data['all_disposals'], 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_book_value': data['total_book_value'],
        'total_disposal_value': data['total_disposal_value'],
        'total_disposal_cost': data['total_disposal_cost'],
        'start_date': start_date,
        'end_date': end_date,
        'company': company,
        'is_super_admin': is_super_admin,
        'showing_all_companies': not company and is_super_admin,
    }
    
    return render(request, 'assets/report_disposal.html', context)


@login_required
def report_by_category(request):
    """Asset by Category Report"""
    from .reports import AssetByCategoryReport
    from django.db.models import Count, Sum, Q
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = AssetByCategoryReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        # Company-wise category summary
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        
        company_data = []
        for comp in companies:
            categories = AssetCategory.objects.filter(company=comp, is_deleted=False).annotate(
                asset_count=Count('assets', filter=Q(assets__is_deleted=False)),
                total_value=Sum('assets__purchase_price', filter=Q(assets__is_deleted=False))
            ).filter(asset_count__gt=0).order_by('-asset_count')
            
            if categories.exists():
                company_data.append({
                    'company': comp,
                    'categories': categories,
                    'total_assets': sum(c.asset_count for c in categories),
                    'total_value': sum(c.total_value or 0 for c in categories),
                    'category_count': categories.count()
                })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = AssetByCategoryReport(company=company)
        data = report.generate()
        
        context = {
            'categories': data['categories'],
            'total_assets': data['total_assets'],
            'total_value': data['total_value'],
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_by_category.html', context)


@login_required
def report_by_location(request):
    """Asset by Location Report"""
    from .reports import AssetByLocationReport
    from django.db.models import Count, Sum, Q
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = AssetByLocationReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        # Company-wise location summary
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        
        company_data = []
        for comp in companies:
            locations = Location.objects.filter(company=comp, is_deleted=False).annotate(
                asset_count=Count('assets', filter=Q(assets__is_deleted=False)),
                total_value=Sum('assets__purchase_price', filter=Q(assets__is_deleted=False))
            ).filter(asset_count__gt=0).order_by('-asset_count')
            
            if locations.exists():
                company_data.append({
                    'company': comp,
                    'locations': locations,
                    'total_assets': sum(l.asset_count for l in locations),
                    'total_value': sum(l.total_value or 0 for l in locations),
                    'location_count': locations.count()
                })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = AssetByLocationReport(company=company)
        data = report.generate()
        
        context = {
            'locations': data['locations'],
            'total_assets': data['total_assets'],
            'total_value': data['total_value'],
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_by_location.html', context)


@login_required
def report_depreciation(request):
    """Depreciation Schedule Report"""
    if not _can_access_financial(request):
        messages.error(request, 'You do not have permission to access depreciation reports.')
        return redirect('assets:dashboard')
    from .reports import DepreciationScheduleReport
    from django.db.models import Sum, Count
    from assets.utils import calculate_current_book_value

    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = DepreciationScheduleReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        from core.models import Company
        from decimal import Decimal
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        
        company_data = []
        for comp in companies:
            assets = Asset.objects.filter(company=comp, is_deleted=False, purchase_price__isnull=False)
            if assets.exists():
                total_purchase = assets.aggregate(total=Sum('purchase_price'))['total'] or 0
                total_current = sum(calculate_current_book_value(asset) for asset in assets)
                total_depreciation = Decimal(str(total_purchase)) - Decimal(str(total_current))
                
                company_data.append({
                    'company': comp,
                    'asset_count': assets.count(),
                    'total_original_value': total_purchase,
                    'total_book_value': total_current,
                    'total_accumulated_depreciation': total_depreciation,
                    'depreciation_percentage': round((float(total_depreciation) / float(total_purchase) * 100), 2) if total_purchase > 0 else 0
                })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = DepreciationScheduleReport(company=company)
        data = report.generate()
        
        # Pagination
        paginator = Paginator(data['schedule'], 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'total_original_value': data['total_original_value'],
            'total_accumulated_depreciation': data['total_accumulated_depreciation'],
            'total_book_value': data['total_book_value'],
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_depreciation.html', context)


@login_required
def report_warranty(request):
    """Warranty Report"""
    from .reports import WarrantyReport
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = WarrantyReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        
        company_data = []
        for comp in companies:
            assets = Asset.objects.filter(company=comp, is_deleted=False)
            under_warranty = assets.filter(warranty_start_date__lte=today, warranty_end_date__gte=today).count()
            expiring_30 = assets.filter(warranty_end_date__gte=today, warranty_end_date__lte=thirty_days).count()
            expired = assets.filter(warranty_end_date__lt=today).count()
            no_warranty = assets.filter(warranty_end_date__isnull=True).count()
            warranty_value = assets.filter(warranty_start_date__lte=today, warranty_end_date__gte=today).aggregate(
                total=Sum('purchase_price')
            )['total'] or 0
            
            # Include company even if all values are 0 to show complete picture
            company_data.append({
                'company': comp,
                'under_warranty': under_warranty,
                'expiring_30_days': expiring_30,
                'expired': expired,
                'no_warranty': no_warranty,
                'total_warranty_value': warranty_value,
                'total_assets': assets.count()
            })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = WarrantyReport(company=company)
        data = report.generate()
        
        context = {
            'under_warranty': data['under_warranty'],
            'expiring_30_days': data['expiring_30_days'],
            'expiring_60_days': data['expiring_60_days'],
            'expiring_90_days': data['expiring_90_days'],
            'expired': data['expired'][:20],  # Limit to 20
            'no_warranty': data['no_warranty'][:20],  # Limit to 20
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_warranty.html', context)


@login_required
def report_amc(request):
    """AMC Report"""
    from .reports import AMCReport
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = AMCReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        
        company_data = []
        for comp in companies:
            assets = Asset.objects.filter(company=comp, is_deleted=False)
            under_amc = assets.filter(amc_start_date__lte=today, amc_end_date__gte=today).count()
            expiring_30 = assets.filter(amc_end_date__gte=today, amc_end_date__lte=thirty_days).count()
            expired = assets.filter(amc_end_date__lt=today).count()
            no_amc = assets.filter(amc_end_date__isnull=True).count()
            total_amc_cost = assets.filter(amc_start_date__lte=today, amc_end_date__gte=today).aggregate(
                total=Sum('amc_cost')
            )['total'] or 0
            
            if under_amc > 0 or expiring_30 > 0 or expired > 0:
                company_data.append({
                    'company': comp,
                    'under_amc': under_amc,
                    'expiring_30_days': expiring_30,
                    'expired': expired,
                    'no_amc': no_amc,
                    'total_amc_cost': total_amc_cost,
                    'total_assets': assets.count()
                })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = AMCReport(company=company)
        data = report.generate()
        
        context = {
            'under_amc': data['under_amc'],
            'expiring_30_days': data['expiring_30_days'],
            'expiring_60_days': data['expiring_60_days'],
            'expiring_90_days': data['expiring_90_days'],
            'expired': data['expired'][:20],  # Limit to 20
            'no_amc': data['no_amc'][:20],  # Limit to 20
            'total_amc_cost': data['total_amc_cost'],
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_amc.html', context)


@login_required
def report_assignment(request):
    """Asset Assignment Report"""
    from .reports import AssignmentReport
    from django.db.models import Count, Q
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    if request.GET.get('export') == 'excel':
        report = AssignmentReport(company=company)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        
        company_data = []
        for comp in companies:
            assets = Asset.objects.filter(company=comp, is_deleted=False)
            assigned = assets.filter(assigned_to__isnull=False).count()
            unassigned = assets.filter(assigned_to__isnull=True).count()
            custodian_assets = assets.filter(custodian__isnull=False).count()
            
            company_data.append({
                'company': comp,
                'total_assigned': assigned,
                'total_unassigned': unassigned,
                'custodian_assets': custodian_assets,
                'total_assets': assets.count(),
                'assignment_percentage': round((assigned / assets.count() * 100), 2) if assets.count() > 0 else 0
            })
        
        context = {
            'company_data': company_data,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = AssignmentReport(company=company)
        data = report.generate()
        
        context = {
            'assignments': data['assignments'],
            'total_assigned': data['total_assigned'],
            'total_unassigned': data['total_unassigned'],
            'unassigned_assets': data['unassigned_assets'][:20],  # Limit to 20
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_assignment.html', context)


@login_required
def report_movement(request):
    """Asset Movement Report"""
    from .reports import MovementReport
    from datetime import datetime, timedelta
    from django.db.models import Count, Q
    
    company = getattr(request, 'current_company', None)
    is_super_admin = getattr(request, 'is_super_admin', False)
    
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = datetime.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if request.GET.get('export') == 'excel':
        report = MovementReport(company=company, start_date=start_date, end_date=end_date)
        return report.export_to_excel()
    
    # For super admin, show company-wise breakdown
    if is_super_admin and not company:
        from core.models import Company
        companies = Company.objects.filter(is_deleted=False, is_active=True)
        
        company_data = []
        for comp in companies:
            movements = AssetHistory.objects.filter(
                asset__company=comp,
                action_date__date__gte=start_date,
                action_date__date__lte=end_date
            )
            
            if movements.exists():
                company_data.append({
                    'company': comp,
                    'total_movements': movements.count(),
                    'location_changes': movements.filter(action_type='LOCATION_CHANGED').count(),
                    'assignments': movements.filter(action_type='ASSIGNED').count(),
                    'status_changes': movements.filter(action_type='STATUS_CHANGED').count(),
                    'transfers': movements.filter(action_type='TRANSFERRED').count(),
                })
        
        context = {
            'company_data': company_data,
            'start_date': start_date,
            'end_date': end_date,
            'is_super_admin': is_super_admin,
            'showing_all_companies': True,
        }
    else:
        # Regular report for specific company
        report = MovementReport(company=company, start_date=start_date, end_date=end_date)
        data = report.generate()
        
        # Pagination
        paginator = Paginator(data['all_movements'], 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'total_movements': data['total_movements'],
            'location_changes': data['location_changes'],
            'assignments': data['assignments'],
            'transfers': data['transfers'],
            'start_date': start_date,
            'end_date': end_date,
            'company': company,
            'is_super_admin': is_super_admin,
            'showing_all_companies': False,
        }
    
    return render(request, 'assets/report_movement.html', context)
