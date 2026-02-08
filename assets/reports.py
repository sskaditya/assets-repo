"""
Comprehensive Reporting & Analytics Module
Provides predefined and customizable report templates
"""

from django.db.models import Count, Sum, Avg, Q, F, DecimalField, Value
from django.db.models.functions import Coalesce, TruncMonth
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from .models import Asset, AssetHistory, AssetTransfer, AssetDisposal
from users.models import Department, Location
from core.models import Company


class ReportGenerator:
    """Base class for generating reports"""
    
    def __init__(self, company=None, start_date=None, end_date=None):
        self.company = company
        self.start_date = start_date or (timezone.now() - timedelta(days=30)).date()
        self.end_date = end_date or timezone.now().date()
    
    def get_base_queryset(self):
        """Get base asset queryset filtered by company"""
        qs = Asset.objects.filter(is_deleted=False)
        if self.company:
            qs = qs.filter(company=self.company)
        return qs
    
    def create_excel_workbook(self, title="Report"):
        """Create a new Excel workbook with styling"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        return wb, ws
    
    def style_header(self, ws, row=1, cols=None):
        """Apply header styling to worksheet"""
        header_fill = PatternFill(start_color="C17845", end_color="C17845", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        if cols:
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


class AssetSummaryReport(ReportGenerator):
    """Asset Summary Report - Overview of all assets"""
    
    def generate(self):
        """Generate asset summary data"""
        assets = self.get_base_queryset()
        
        summary = {
            'total_assets': assets.count(),
            'by_status': assets.values('status').annotate(count=Count('id')).order_by('-count'),
            'by_condition': assets.values('condition').annotate(count=Count('id')).order_by('-count'),
            'by_category': assets.values('category__name').annotate(count=Count('id')).order_by('-count'),
            'by_location': assets.values('location__name').annotate(count=Count('id')).order_by('-count'),
            'by_department': assets.values('department__name').annotate(count=Count('id')).order_by('-count'),
            'critical_assets': assets.filter(is_critical=True).count(),
            'total_value': assets.aggregate(total=Sum('purchase_price'))['total'] or 0,
            'avg_value': assets.aggregate(avg=Avg('purchase_price'))['avg'] or 0,
        }
        
        return summary
    
    def export_to_excel(self):
        """Export asset summary to Excel"""
        wb, ws = self.create_excel_workbook("Asset Summary")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = 'ASSET SUMMARY REPORT'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date range
        ws['A2'] = f'Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        if self.company:
            ws['A3'] = f'Company: {self.company.name}'
        
        # Summary data
        summary = self.generate()
        row = 5
        
        # Overall stats
        ws[f'A{row}'] = 'OVERALL STATISTICS'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Total Assets'
        ws[f'B{row}'] = summary['total_assets']
        row += 1
        
        ws[f'A{row}'] = 'Critical Assets'
        ws[f'B{row}'] = summary['critical_assets']
        row += 1
        
        ws[f'A{row}'] = 'Total Value'
        ws[f'B{row}'] = float(summary['total_value'])
        row += 1
        
        ws[f'A{row}'] = 'Average Value'
        ws[f'B{row}'] = float(summary['avg_value'])
        row += 2
        
        # By Status
        ws[f'A{row}'] = 'BY STATUS'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for item in summary['by_status']:
            ws[f'A{row}'] = item['status']
            ws[f'B{row}'] = item['count']
            row += 1
        row += 1
        
        # By Category
        ws[f'A{row}'] = 'BY CATEGORY'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for item in summary['by_category']:
            ws[f'A{row}'] = item['category__name'] or 'Uncategorized'
            ws[f'B{row}'] = item['count']
            row += 1
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=asset_summary_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class AssetListReport(ReportGenerator):
    """Detailed Asset List Report"""
    
    def generate(self, filters=None):
        """Generate filtered asset list"""
        assets = self.get_base_queryset()
        
        if filters:
            if filters.get('status'):
                assets = assets.filter(status=filters['status'])
            if filters.get('category'):
                assets = assets.filter(category_id=filters['category'])
            if filters.get('location'):
                assets = assets.filter(location_id=filters['location'])
            if filters.get('department'):
                assets = assets.filter(department_id=filters['department'])
            if filters.get('condition'):
                assets = assets.filter(condition=filters['condition'])
        
        return assets.select_related(
            'category', 'asset_type', 'location', 'department', 'assigned_to', 'vendor'
        ).order_by('asset_tag')
    
    def export_to_excel(self, filters=None):
        """Export asset list to Excel"""
        wb, ws = self.create_excel_workbook("Asset List")
        
        # Headers
        headers = [
            'Asset Tag', 'Name', 'Category', 'Type', 'Status', 'Condition',
            'Location', 'Department', 'Assigned To', 'Serial Number',
            'Make', 'Model', 'Purchase Date', 'Purchase Price', 'Vendor'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self.style_header(ws, row=1, cols=len(headers))
        
        # Data
        assets = self.generate(filters)
        for row, asset in enumerate(assets, 2):
            ws.cell(row=row, column=1, value=asset.asset_tag)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
            ws.cell(row=row, column=4, value=asset.asset_type.name if asset.asset_type else '')
            ws.cell(row=row, column=5, value=asset.status)
            ws.cell(row=row, column=6, value=asset.condition)
            ws.cell(row=row, column=7, value=asset.location.name if asset.location else '')
            ws.cell(row=row, column=8, value=asset.department.name if asset.department else '')
            ws.cell(row=row, column=9, value=asset.assigned_to.get_full_name() if asset.assigned_to else '')
            ws.cell(row=row, column=10, value=asset.serial_number or '')
            ws.cell(row=row, column=11, value=asset.make or '')
            ws.cell(row=row, column=12, value=asset.model or '')
            ws.cell(row=row, column=13, value=asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '')
            ws.cell(row=row, column=14, value=float(asset.purchase_price) if asset.purchase_price else 0)
            ws.cell(row=row, column=15, value=asset.vendor.name if asset.vendor else '')
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=asset_list_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class FinancialReport(ReportGenerator):
    """Financial Report - Asset values, depreciation"""
    
    def generate(self):
        """Generate financial data"""
        from .utils import calculate_current_book_value
        
        assets = self.get_base_queryset().filter(purchase_price__isnull=False)
        
        total_purchase_value = assets.aggregate(total=Sum('purchase_price'))['total'] or 0
        total_book_value = Decimal('0')
        total_depreciation = Decimal('0')
        
        asset_financials = []
        for asset in assets:
            book_value = calculate_current_book_value(asset)
            depreciation = asset.purchase_price - book_value
            total_book_value += book_value
            total_depreciation += depreciation
            
            asset_financials.append({
                'asset': asset,
                'purchase_price': asset.purchase_price,
                'book_value': book_value,
                'depreciation': depreciation,
                'depreciation_percent': (depreciation / asset.purchase_price * 100) if asset.purchase_price else 0
            })
        
        return {
            'total_purchase_value': total_purchase_value,
            'total_book_value': total_book_value,
            'total_depreciation': total_depreciation,
            'asset_financials': asset_financials,
        }
    
    def export_to_excel(self):
        """Export financial report to Excel"""
        wb, ws = self.create_excel_workbook("Financial Report")
        
        # Headers
        headers = [
            'Asset Tag', 'Name', 'Category', 'Purchase Date', 'Purchase Price',
            'Depreciation Rate', 'Useful Life', 'Current Book Value',
            'Total Depreciation', 'Depreciation %'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self.style_header(ws, row=1, cols=len(headers))
        
        # Data
        data = self.generate()
        row = 2
        for item in data['asset_financials']:
            asset = item['asset']
            ws.cell(row=row, column=1, value=asset.asset_tag)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
            ws.cell(row=row, column=4, value=asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '')
            ws.cell(row=row, column=5, value=float(item['purchase_price']))
            ws.cell(row=row, column=6, value=float(asset.depreciation_rate) if asset.depreciation_rate else 0)
            ws.cell(row=row, column=7, value=asset.useful_life_years or 0)
            ws.cell(row=row, column=8, value=float(item['book_value']))
            ws.cell(row=row, column=9, value=float(item['depreciation']))
            ws.cell(row=row, column=10, value=f"{item['depreciation_percent']:.2f}%")
            row += 1
        
        # Summary
        row += 2
        ws.cell(row=row, column=1, value='SUMMARY').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value='Total Purchase Value')
        ws.cell(row=row, column=2, value=float(data['total_purchase_value']))
        row += 1
        ws.cell(row=row, column=1, value='Total Current Book Value')
        ws.cell(row=row, column=2, value=float(data['total_book_value']))
        row += 1
        ws.cell(row=row, column=1, value='Total Depreciation')
        ws.cell(row=row, column=2, value=float(data['total_depreciation']))
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=financial_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class MaintenanceReport(ReportGenerator):
    """Maintenance Report - Warranty, AMC, maintenance schedule"""
    
    def generate(self):
        """Generate maintenance data"""
        assets = self.get_base_queryset()
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        
        return {
            'warranty_expiring': assets.filter(
                warranty_end_date__gte=today,
                warranty_end_date__lte=thirty_days
            ).order_by('warranty_end_date'),
            'warranty_expired': assets.filter(
                warranty_end_date__lt=today
            ).order_by('-warranty_end_date'),
            'amc_expiring': assets.filter(
                amc_end_date__gte=today,
                amc_end_date__lte=thirty_days
            ).order_by('amc_end_date'),
            'amc_expired': assets.filter(
                amc_end_date__lt=today
            ).order_by('-amc_end_date'),
            'under_warranty': assets.filter(
                warranty_start_date__lte=today,
                warranty_end_date__gte=today
            ).count(),
            'under_amc': assets.filter(
                amc_start_date__lte=today,
                amc_end_date__gte=today
            ).count(),
        }
    
    def export_to_excel(self):
        """Export maintenance report to Excel"""
        wb, ws = self.create_excel_workbook("Maintenance Report")
        
        data = self.generate()
        row = 1
        
        # Warranty Expiring
        ws.cell(row=row, column=1, value='WARRANTY EXPIRING (Next 30 Days)').font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Asset Tag', 'Name', 'Category', 'Warranty Start', 'Warranty End', 'Days Remaining']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.style_header(ws, row=row, cols=len(headers))
        row += 1
        
        for asset in data['warranty_expiring']:
            days_remaining = (asset.warranty_end_date - timezone.now().date()).days
            ws.cell(row=row, column=1, value=asset.asset_tag)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
            ws.cell(row=row, column=4, value=asset.warranty_start_date.strftime('%Y-%m-%d') if asset.warranty_start_date else '')
            ws.cell(row=row, column=5, value=asset.warranty_end_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=6, value=days_remaining)
            row += 1
        
        row += 2
        
        # AMC Expiring
        ws.cell(row=row, column=1, value='AMC EXPIRING (Next 30 Days)').font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Asset Tag', 'Name', 'Category', 'AMC Vendor', 'AMC Start', 'AMC End', 'Days Remaining', 'AMC Cost']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self.style_header(ws, row=row, cols=len(headers))
        row += 1
        
        for asset in data['amc_expiring']:
            days_remaining = (asset.amc_end_date - timezone.now().date()).days
            ws.cell(row=row, column=1, value=asset.asset_tag)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
            ws.cell(row=row, column=4, value=asset.amc_vendor.name if asset.amc_vendor else '')
            ws.cell(row=row, column=5, value=asset.amc_start_date.strftime('%Y-%m-%d') if asset.amc_start_date else '')
            ws.cell(row=row, column=6, value=asset.amc_end_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=days_remaining)
            ws.cell(row=row, column=8, value=float(asset.amc_cost) if asset.amc_cost else 0)
            row += 1
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=maintenance_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class TransferReport(ReportGenerator):
    """Asset Transfer Report"""
    
    def generate(self):
        """Generate transfer data"""
        from .models import AssetTransfer
        
        transfers = AssetTransfer.objects.filter(
            requested_date__date__gte=self.start_date,
            requested_date__date__lte=self.end_date
        )
        
        if self.company:
            transfers = transfers.filter(asset__company=self.company)
        
        return {
            'all_transfers': transfers.select_related(
                'asset', 'from_location', 'to_location', 'requested_by', 'approved_by'
            ).order_by('-requested_date'),
            'by_status': transfers.values('status').annotate(count=Count('id')),
            'pending_count': transfers.filter(status='PENDING').count(),
            'approved_count': transfers.filter(status='APPROVED').count(),
            'completed_count': transfers.filter(status='COMPLETED').count(),
            'rejected_count': transfers.filter(status='REJECTED').count(),
        }
    
    def export_to_excel(self):
        """Export transfer report to Excel"""
        wb, ws = self.create_excel_workbook("Transfer Report")
        
        # Headers
        headers = [
            'Transfer Number', 'Asset Tag', 'Asset Name', 'Status',
            'From Location', 'To Location', 'From User', 'To User',
            'Requested By', 'Requested Date', 'Approved By', 'Approval Date', 'Reason'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self.style_header(ws, row=1, cols=len(headers))
        
        # Data
        data = self.generate()
        row = 2
        for transfer in data['all_transfers']:
            ws.cell(row=row, column=1, value=transfer.transfer_number)
            ws.cell(row=row, column=2, value=transfer.asset.asset_tag)
            ws.cell(row=row, column=3, value=transfer.asset.name)
            ws.cell(row=row, column=4, value=transfer.status)
            ws.cell(row=row, column=5, value=transfer.from_location.name if transfer.from_location else '')
            ws.cell(row=row, column=6, value=transfer.to_location.name if transfer.to_location else '')
            ws.cell(row=row, column=7, value=transfer.from_user.get_full_name() if transfer.from_user else '')
            ws.cell(row=row, column=8, value=transfer.to_user.get_full_name() if transfer.to_user else '')
            ws.cell(row=row, column=9, value=transfer.requested_by.get_full_name() if transfer.requested_by else '')
            ws.cell(row=row, column=10, value=transfer.requested_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=11, value=transfer.approved_by.get_full_name() if transfer.approved_by else '')
            ws.cell(row=row, column=12, value=transfer.approval_date.strftime('%Y-%m-%d %H:%M') if transfer.approval_date else '')
            ws.cell(row=row, column=13, value=transfer.reason)
            row += 1
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=transfer_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class DisposalReport(ReportGenerator):
    """Asset Disposal Report"""
    
    def generate(self):
        """Generate disposal data"""
        from .models import AssetDisposal
        
        disposals = AssetDisposal.objects.filter(
            requested_date__date__gte=self.start_date,
            requested_date__date__lte=self.end_date
        )
        
        if self.company:
            disposals = disposals.filter(asset__company=self.company)
        
        return {
            'all_disposals': disposals.select_related(
                'asset', 'requested_by', 'approved_by'
            ).order_by('-requested_date'),
            'by_status': disposals.values('status').annotate(count=Count('id')),
            'by_method': disposals.values('disposal_method').annotate(count=Count('id')),
            'total_book_value': disposals.aggregate(total=Sum('current_book_value'))['total'] or 0,
            'total_disposal_value': disposals.aggregate(total=Sum('disposal_value'))['total'] or 0,
            'total_disposal_cost': disposals.aggregate(total=Sum('disposal_cost'))['total'] or 0,
        }
    
    def export_to_excel(self):
        """Export disposal report to Excel"""
        wb, ws = self.create_excel_workbook("Disposal Report")
        
        # Headers
        headers = [
            'Disposal Number', 'Asset Tag', 'Asset Name', 'Status', 'Disposal Method',
            'Book Value', 'Disposal Value', 'Disposal Cost', 'Gain/Loss',
            'Requested By', 'Requested Date', 'Approved By', 'Approval Date', 'Reason'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self.style_header(ws, row=1, cols=len(headers))
        
        # Data
        data = self.generate()
        row = 2
        for disposal in data['all_disposals']:
            gain_loss = (disposal.disposal_value - disposal.disposal_cost - (disposal.current_book_value or 0))
            ws.cell(row=row, column=1, value=disposal.disposal_number)
            ws.cell(row=row, column=2, value=disposal.asset.asset_tag)
            ws.cell(row=row, column=3, value=disposal.asset.name)
            ws.cell(row=row, column=4, value=disposal.status)
            ws.cell(row=row, column=5, value=disposal.disposal_method)
            ws.cell(row=row, column=6, value=float(disposal.current_book_value) if disposal.current_book_value else 0)
            ws.cell(row=row, column=7, value=float(disposal.disposal_value))
            ws.cell(row=row, column=8, value=float(disposal.disposal_cost))
            ws.cell(row=row, column=9, value=float(gain_loss))
            ws.cell(row=row, column=10, value=disposal.requested_by.get_full_name() if disposal.requested_by else '')
            ws.cell(row=row, column=11, value=disposal.requested_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=12, value=disposal.approved_by.get_full_name() if disposal.approved_by else '')
            ws.cell(row=row, column=13, value=disposal.approval_date.strftime('%Y-%m-%d %H:%M') if disposal.approval_date else '')
            ws.cell(row=row, column=14, value=disposal.reason)
            row += 1
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=disposal_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class AssetByCategoryReport(ReportGenerator):
    """Asset by Category Report"""
    
    def generate(self):
        """Generate asset breakdown by category"""
        from .models import AssetCategory
        
        assets = self.get_base_queryset()
        
        if self.company:
            categories = AssetCategory.objects.filter(company=self.company, is_deleted=False)
        else:
            categories = AssetCategory.objects.filter(is_deleted=False)
        
        category_data = []
        for category in categories:
            category_assets = assets.filter(category=category)
            total_value = category_assets.aggregate(total=Sum('purchase_price'))['total'] or 0
            
            category_data.append({
                'category': category,
                'count': category_assets.count(),
                'total_value': total_value,
                'by_status': category_assets.values('status').annotate(count=Count('id')),
                'assets': category_assets.select_related('asset_type', 'location', 'department')
            })
        
        return {
            'categories': category_data,
            'total_assets': assets.count(),
            'total_value': assets.aggregate(total=Sum('purchase_price'))['total'] or 0,
        }
    
    def export_to_excel(self):
        """Export asset by category to Excel"""
        wb, ws = self.create_excel_workbook("Assets by Category")
        
        data = self.generate()
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'ASSETS BY CATEGORY REPORT'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        for cat_data in data['categories']:
            # Category header
            ws[f'A{row}'] = f'{cat_data["category"].name}'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = f'Total Assets: {cat_data["count"]}'
            ws[f'C{row}'] = f'Total Value: {float(cat_data["total_value"])}'
            row += 1
            
            # Asset details
            headers = ['Asset Tag', 'Name', 'Type', 'Status', 'Location', 'Purchase Price']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self.style_header(ws, row=row, cols=len(headers))
            row += 1
            
            for asset in cat_data['assets']:
                ws.cell(row=row, column=1, value=asset.asset_tag)
                ws.cell(row=row, column=2, value=asset.name)
                ws.cell(row=row, column=3, value=asset.asset_type.name if asset.asset_type else '')
                ws.cell(row=row, column=4, value=asset.status)
                ws.cell(row=row, column=5, value=asset.location.name if asset.location else '')
                ws.cell(row=row, column=6, value=float(asset.purchase_price) if asset.purchase_price else 0)
                row += 1
            
            row += 2
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=assets_by_category_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class AssetByLocationReport(ReportGenerator):
    """Asset by Location Report"""
    
    def generate(self):
        """Generate asset breakdown by location"""
        assets = self.get_base_queryset()
        
        if self.company:
            locations = Location.objects.filter(company=self.company, is_deleted=False)
        else:
            locations = Location.objects.all()
        
        location_data = []
        for location in locations:
            location_assets = assets.filter(location=location)
            total_value = location_assets.aggregate(total=Sum('purchase_price'))['total'] or 0
            
            location_data.append({
                'location': location,
                'count': location_assets.count(),
                'total_value': total_value,
                'by_status': location_assets.values('status').annotate(count=Count('id')),
                'by_category': location_assets.values('category__name').annotate(count=Count('id')),
                'assets': location_assets.select_related('category', 'asset_type', 'department')
            })
        
        # Unassigned assets
        unassigned = assets.filter(location__isnull=True)
        if unassigned.exists():
            location_data.append({
                'location': None,
                'count': unassigned.count(),
                'total_value': unassigned.aggregate(total=Sum('purchase_price'))['total'] or 0,
                'by_status': unassigned.values('status').annotate(count=Count('id')),
                'by_category': unassigned.values('category__name').annotate(count=Count('id')),
                'assets': unassigned.select_related('category', 'asset_type', 'department')
            })
        
        return {
            'locations': location_data,
            'total_assets': assets.count(),
            'total_value': assets.aggregate(total=Sum('purchase_price'))['total'] or 0,
        }
    
    def export_to_excel(self):
        """Export asset by location to Excel"""
        wb, ws = self.create_excel_workbook("Assets by Location")
        
        data = self.generate()
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'ASSETS BY LOCATION REPORT'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        for loc_data in data['locations']:
            # Location header
            location_name = loc_data['location'].name if loc_data['location'] else 'Unassigned'
            ws[f'A{row}'] = f'{location_name}'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = f'Total Assets: {loc_data["count"]}'
            ws[f'C{row}'] = f'Total Value: {float(loc_data["total_value"])}'
            row += 1
            
            # Asset details
            headers = ['Asset Tag', 'Name', 'Category', 'Status', 'Department', 'Purchase Price']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self.style_header(ws, row=row, cols=len(headers))
            row += 1
            
            for asset in loc_data['assets']:
                ws.cell(row=row, column=1, value=asset.asset_tag)
                ws.cell(row=row, column=2, value=asset.name)
                ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
                ws.cell(row=row, column=4, value=asset.status)
                ws.cell(row=row, column=5, value=asset.department.name if asset.department else '')
                ws.cell(row=row, column=6, value=float(asset.purchase_price) if asset.purchase_price else 0)
                row += 1
            
            row += 2
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=assets_by_location_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class DepreciationScheduleReport(ReportGenerator):
    """Depreciation Schedule Report"""
    
    def generate(self):
        """Generate depreciation schedule"""
        from .utils import calculate_current_book_value
        
        assets = self.get_base_queryset().filter(
            purchase_price__isnull=False,
            depreciation_rate__isnull=False
        )
        
        schedule = []
        total_original_value = Decimal('0')
        total_accumulated_depreciation = Decimal('0')
        total_book_value = Decimal('0')
        
        for asset in assets:
            book_value = calculate_current_book_value(asset)
            accumulated_depreciation = asset.purchase_price - book_value
            
            # Calculate yearly depreciation
            if asset.purchase_date:
                years_owned = (timezone.now().date() - asset.purchase_date).days / 365.25
            else:
                years_owned = 0
            
            annual_depreciation = asset.purchase_price * (asset.depreciation_rate / 100) if asset.depreciation_rate else 0
            
            schedule.append({
                'asset': asset,
                'original_value': asset.purchase_price,
                'depreciation_rate': asset.depreciation_rate,
                'useful_life': asset.useful_life_years,
                'years_owned': round(years_owned, 2),
                'annual_depreciation': annual_depreciation,
                'accumulated_depreciation': accumulated_depreciation,
                'book_value': book_value,
                'salvage_value': asset.salvage_value or 0,
            })
            
            total_original_value += asset.purchase_price
            total_accumulated_depreciation += accumulated_depreciation
            total_book_value += book_value
        
        return {
            'schedule': schedule,
            'total_original_value': total_original_value,
            'total_accumulated_depreciation': total_accumulated_depreciation,
            'total_book_value': total_book_value,
        }
    
    def export_to_excel(self):
        """Export depreciation schedule to Excel"""
        wb, ws = self.create_excel_workbook("Depreciation Schedule")
        
        # Headers
        headers = [
            'Asset Tag', 'Name', 'Category', 'Purchase Date', 'Original Value',
            'Depreciation Rate (%)', 'Useful Life (Years)', 'Years Owned',
            'Annual Depreciation', 'Accumulated Depreciation', 'Current Book Value', 'Salvage Value'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self.style_header(ws, row=1, cols=len(headers))
        
        # Data
        data = self.generate()
        row = 2
        for item in data['schedule']:
            asset = item['asset']
            ws.cell(row=row, column=1, value=asset.asset_tag)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
            ws.cell(row=row, column=4, value=asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '')
            ws.cell(row=row, column=5, value=float(item['original_value']))
            ws.cell(row=row, column=6, value=float(item['depreciation_rate']) if item['depreciation_rate'] else 0)
            ws.cell(row=row, column=7, value=item['useful_life'] or 0)
            ws.cell(row=row, column=8, value=item['years_owned'])
            ws.cell(row=row, column=9, value=float(item['annual_depreciation']))
            ws.cell(row=row, column=10, value=float(item['accumulated_depreciation']))
            ws.cell(row=row, column=11, value=float(item['book_value']))
            ws.cell(row=row, column=12, value=float(item['salvage_value']))
            row += 1
        
        # Summary
        row += 2
        ws.cell(row=row, column=1, value='SUMMARY').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value='Total Original Value')
        ws.cell(row=row, column=2, value=float(data['total_original_value']))
        row += 1
        ws.cell(row=row, column=1, value='Total Accumulated Depreciation')
        ws.cell(row=row, column=2, value=float(data['total_accumulated_depreciation']))
        row += 1
        ws.cell(row=row, column=1, value='Total Current Book Value')
        ws.cell(row=row, column=2, value=float(data['total_book_value']))
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=depreciation_schedule_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class WarrantyReport(ReportGenerator):
    """Warranty Report"""
    
    def generate(self):
        """Generate warranty data"""
        assets = self.get_base_queryset()
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        sixty_days = today + timedelta(days=60)
        ninety_days = today + timedelta(days=90)
        
        return {
            'under_warranty': assets.filter(
                warranty_start_date__lte=today,
                warranty_end_date__gte=today
            ).select_related('category', 'location', 'vendor').order_by('warranty_end_date'),
            'expiring_30_days': assets.filter(
                warranty_end_date__gte=today,
                warranty_end_date__lte=thirty_days
            ).select_related('category', 'location', 'vendor').order_by('warranty_end_date'),
            'expiring_60_days': assets.filter(
                warranty_end_date__gt=thirty_days,
                warranty_end_date__lte=sixty_days
            ).select_related('category', 'location', 'vendor').order_by('warranty_end_date'),
            'expiring_90_days': assets.filter(
                warranty_end_date__gt=sixty_days,
                warranty_end_date__lte=ninety_days
            ).select_related('category', 'location', 'vendor').order_by('warranty_end_date'),
            'expired': assets.filter(
                warranty_end_date__lt=today
            ).select_related('category', 'location', 'vendor').order_by('-warranty_end_date'),
            'no_warranty': assets.filter(
                warranty_end_date__isnull=True
            ).select_related('category', 'location').order_by('asset_tag'),
        }
    
    def export_to_excel(self):
        """Export warranty report to Excel"""
        wb, ws = self.create_excel_workbook("Warranty Report")
        
        data = self.generate()
        row = 1
        
        sections = [
            ('UNDER WARRANTY', data['under_warranty']),
            ('EXPIRING IN 30 DAYS', data['expiring_30_days']),
            ('EXPIRING IN 60 DAYS', data['expiring_60_days']),
            ('EXPIRING IN 90 DAYS', data['expiring_90_days']),
            ('EXPIRED', data['expired']),
            ('NO WARRANTY', data['no_warranty']),
        ]
        
        for section_title, assets in sections:
            ws.cell(row=row, column=1, value=f'{section_title} ({assets.count()})').font = Font(bold=True, size=12)
            row += 1
            
            if assets.exists():
                headers = ['Asset Tag', 'Name', 'Category', 'Location', 'Vendor', 'Warranty Start', 'Warranty End', 'Days Remaining']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                self.style_header(ws, row=row, cols=len(headers))
                row += 1
                
                for asset in assets:
                    days_remaining = (asset.warranty_end_date - today).days if asset.warranty_end_date else 'N/A'
                    ws.cell(row=row, column=1, value=asset.asset_tag)
                    ws.cell(row=row, column=2, value=asset.name)
                    ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
                    ws.cell(row=row, column=4, value=asset.location.name if asset.location else '')
                    ws.cell(row=row, column=5, value=asset.vendor.name if asset.vendor else '')
                    ws.cell(row=row, column=6, value=asset.warranty_start_date.strftime('%Y-%m-%d') if asset.warranty_start_date else '')
                    ws.cell(row=row, column=7, value=asset.warranty_end_date.strftime('%Y-%m-%d') if asset.warranty_end_date else '')
                    ws.cell(row=row, column=8, value=days_remaining if days_remaining != 'N/A' else 'N/A')
                    row += 1
            
            row += 2
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=warranty_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class AMCReport(ReportGenerator):
    """AMC (Annual Maintenance Contract) Report"""
    
    def generate(self):
        """Generate AMC data"""
        assets = self.get_base_queryset()
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        sixty_days = today + timedelta(days=60)
        ninety_days = today + timedelta(days=90)
        
        return {
            'under_amc': assets.filter(
                amc_start_date__lte=today,
                amc_end_date__gte=today
            ).select_related('category', 'location', 'amc_vendor').order_by('amc_end_date'),
            'expiring_30_days': assets.filter(
                amc_end_date__gte=today,
                amc_end_date__lte=thirty_days
            ).select_related('category', 'location', 'amc_vendor').order_by('amc_end_date'),
            'expiring_60_days': assets.filter(
                amc_end_date__gt=thirty_days,
                amc_end_date__lte=sixty_days
            ).select_related('category', 'location', 'amc_vendor').order_by('amc_end_date'),
            'expiring_90_days': assets.filter(
                amc_end_date__gt=sixty_days,
                amc_end_date__lte=ninety_days
            ).select_related('category', 'location', 'amc_vendor').order_by('amc_end_date'),
            'expired': assets.filter(
                amc_end_date__lt=today
            ).select_related('category', 'location', 'amc_vendor').order_by('-amc_end_date'),
            'no_amc': assets.filter(
                amc_end_date__isnull=True
            ).select_related('category', 'location').order_by('asset_tag'),
            'total_amc_cost': assets.filter(
                amc_start_date__lte=today,
                amc_end_date__gte=today
            ).aggregate(total=Sum('amc_cost'))['total'] or 0,
        }
    
    def export_to_excel(self):
        """Export AMC report to Excel"""
        wb, ws = self.create_excel_workbook("AMC Report")
        
        data = self.generate()
        row = 1
        
        sections = [
            ('UNDER AMC', data['under_amc']),
            ('EXPIRING IN 30 DAYS', data['expiring_30_days']),
            ('EXPIRING IN 60 DAYS', data['expiring_60_days']),
            ('EXPIRING IN 90 DAYS', data['expiring_90_days']),
            ('EXPIRED', data['expired']),
            ('NO AMC', data['no_amc']),
        ]
        
        for section_title, assets in sections:
            ws.cell(row=row, column=1, value=f'{section_title} ({assets.count()})').font = Font(bold=True, size=12)
            row += 1
            
            if assets.exists():
                headers = ['Asset Tag', 'Name', 'Category', 'Location', 'AMC Vendor', 'AMC Start', 'AMC End', 'Days Remaining', 'AMC Cost']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                self.style_header(ws, row=row, cols=len(headers))
                row += 1
                
                today = timezone.now().date()
                for asset in assets:
                    days_remaining = (asset.amc_end_date - today).days if asset.amc_end_date else 'N/A'
                    ws.cell(row=row, column=1, value=asset.asset_tag)
                    ws.cell(row=row, column=2, value=asset.name)
                    ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
                    ws.cell(row=row, column=4, value=asset.location.name if asset.location else '')
                    ws.cell(row=row, column=5, value=asset.amc_vendor.name if asset.amc_vendor else '')
                    ws.cell(row=row, column=6, value=asset.amc_start_date.strftime('%Y-%m-%d') if asset.amc_start_date else '')
                    ws.cell(row=row, column=7, value=asset.amc_end_date.strftime('%Y-%m-%d') if asset.amc_end_date else '')
                    ws.cell(row=row, column=8, value=days_remaining if days_remaining != 'N/A' else 'N/A')
                    ws.cell(row=row, column=9, value=float(asset.amc_cost) if asset.amc_cost else 0)
                    row += 1
            
            row += 2
        
        # Summary
        ws.cell(row=row, column=1, value='SUMMARY').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value='Total Active AMC Cost')
        ws.cell(row=row, column=2, value=float(data['total_amc_cost']))
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=amc_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class AssignmentReport(ReportGenerator):
    """Asset Assignment Report"""
    
    def generate(self):
        """Generate asset assignment data"""
        assets = self.get_base_queryset()
        
        from django.contrib.auth.models import User
        
        if self.company:
            users = User.objects.filter(profile__company=self.company).distinct()
        else:
            users = User.objects.all()
        
        assignment_data = []
        for user in users:
            user_assets = assets.filter(assigned_to=user)
            custodian_assets = assets.filter(custodian=user)
            
            if user_assets.exists() or custodian_assets.exists():
                assignment_data.append({
                    'user': user,
                    'assigned_count': user_assets.count(),
                    'custodian_count': custodian_assets.count(),
                    'assigned_assets': user_assets.select_related('category', 'location', 'department'),
                    'custodian_assets': custodian_assets.select_related('category', 'location', 'department'),
                    'total_value_assigned': user_assets.aggregate(total=Sum('purchase_price'))['total'] or 0,
                    'total_value_custodian': custodian_assets.aggregate(total=Sum('purchase_price'))['total'] or 0,
                })
        
        return {
            'assignments': assignment_data,
            'total_assigned': assets.filter(assigned_to__isnull=False).count(),
            'total_unassigned': assets.filter(assigned_to__isnull=True).count(),
            'unassigned_assets': assets.filter(assigned_to__isnull=True).select_related('category', 'location', 'department'),
        }
    
    def export_to_excel(self):
        """Export assignment report to Excel"""
        wb, ws = self.create_excel_workbook("Assignment Report")
        
        data = self.generate()
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'ASSET ASSIGNMENT REPORT'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        for assignment in data['assignments']:
            user = assignment['user']
            ws[f'A{row}'] = f'{user.get_full_name()} ({user.username})'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = f'Assigned Assets: {assignment["assigned_count"]}'
            ws[f'C{row}'] = f'Custodian Assets: {assignment["custodian_count"]}'
            row += 1
            
            # Assigned assets
            if assignment['assigned_assets'].exists():
                ws[f'A{row}'] = 'ASSIGNED ASSETS'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                headers = ['Asset Tag', 'Name', 'Category', 'Location', 'Department', 'Purchase Price']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                self.style_header(ws, row=row, cols=len(headers))
                row += 1
                
                for asset in assignment['assigned_assets']:
                    ws.cell(row=row, column=1, value=asset.asset_tag)
                    ws.cell(row=row, column=2, value=asset.name)
                    ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
                    ws.cell(row=row, column=4, value=asset.location.name if asset.location else '')
                    ws.cell(row=row, column=5, value=asset.department.name if asset.department else '')
                    ws.cell(row=row, column=6, value=float(asset.purchase_price) if asset.purchase_price else 0)
                    row += 1
            
            row += 2
        
        # Unassigned assets
        if data['unassigned_assets'].exists():
            ws[f'A{row}'] = f'UNASSIGNED ASSETS ({data["total_unassigned"]})'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Asset Tag', 'Name', 'Category', 'Location', 'Department', 'Status']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self.style_header(ws, row=row, cols=len(headers))
            row += 1
            
            for asset in data['unassigned_assets']:
                ws.cell(row=row, column=1, value=asset.asset_tag)
                ws.cell(row=row, column=2, value=asset.name)
                ws.cell(row=row, column=3, value=asset.category.name if asset.category else '')
                ws.cell(row=row, column=4, value=asset.location.name if asset.location else '')
                ws.cell(row=row, column=5, value=asset.department.name if asset.department else '')
                ws.cell(row=row, column=6, value=asset.status)
                row += 1
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=assignment_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class MovementReport(ReportGenerator):
    """Asset Movement/History Report"""
    
    def generate(self):
        """Generate asset movement data"""
        history = AssetHistory.objects.filter(
            action_date__date__gte=self.start_date,
            action_date__date__lte=self.end_date
        )
        
        if self.company:
            history = history.filter(asset__company=self.company)
        
        return {
            'all_movements': history.select_related(
                'asset', 'performed_by', 'from_location', 'to_location', 'from_user', 'to_user'
            ).order_by('-action_date'),
            'by_action_type': history.values('action_type').annotate(count=Count('id')),
            'total_movements': history.count(),
            'location_changes': history.filter(action_type='LOCATION_CHANGED').count(),
            'assignments': history.filter(action_type='ASSIGNED').count(),
            'transfers': history.filter(action_type='TRANSFERRED').count(),
        }
    
    def export_to_excel(self):
        """Export movement report to Excel"""
        wb, ws = self.create_excel_workbook("Movement Report")
        
        # Headers
        headers = [
            'Date', 'Asset Tag', 'Asset Name', 'Action Type', 'From Location',
            'To Location', 'From User', 'To User', 'Performed By', 'Remarks'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        self.style_header(ws, row=1, cols=len(headers))
        
        # Data
        data = self.generate()
        row = 2
        for movement in data['all_movements']:
            ws.cell(row=row, column=1, value=movement.action_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=2, value=movement.asset.asset_tag if movement.asset else '')
            ws.cell(row=row, column=3, value=movement.asset.name if movement.asset else '')
            ws.cell(row=row, column=4, value=movement.action_type)
            ws.cell(row=row, column=5, value=movement.from_location.name if movement.from_location else '')
            ws.cell(row=row, column=6, value=movement.to_location.name if movement.to_location else '')
            ws.cell(row=row, column=7, value=movement.from_user.get_full_name() if movement.from_user else '')
            ws.cell(row=row, column=8, value=movement.to_user.get_full_name() if movement.to_user else '')
            ws.cell(row=row, column=9, value=movement.performed_by.get_full_name() if movement.performed_by else '')
            ws.cell(row=row, column=10, value=movement.remarks or '')
            row += 1
        
        # Summary
        row += 2
        ws.cell(row=row, column=1, value='SUMMARY').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value='Total Movements')
        ws.cell(row=row, column=2, value=data['total_movements'])
        row += 1
        ws.cell(row=row, column=1, value='Location Changes')
        ws.cell(row=row, column=2, value=data['location_changes'])
        row += 1
        ws.cell(row=row, column=1, value='Assignments')
        ws.cell(row=row, column=2, value=data['assignments'])
        row += 1
        ws.cell(row=row, column=1, value='Transfers')
        ws.cell(row=row, column=2, value=data['transfers'])
        
        self.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=movement_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
